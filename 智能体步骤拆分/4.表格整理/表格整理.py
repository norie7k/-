# -*- coding: utf-8 -*-
from pathlib import Path
import json, sys, shutil
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ========= 路径（按需修改） =========
INPUT_FILE  = r"E:\项目\玩家社群分析智能体\4.表格整理\表格整理.json"
EXCEL_FILE  = r"E:\项目\玩家社群分析智能体\5.最终输出文件\群聊数据_按CA分表.xlsx"
# ===================================

# 基础配置
ENCODING         = "utf-8-sig"
TIME_FORMAT      = "%Y-%m-%d %H:%M:%S"
DO_BACKUP        = True
ALLOW_DUPLICATE  = False

# CA -> 中文 Sheet
CA_TO_SHEET = {1:"体验反馈", 2:"疑惑询问", 3:"建议灵感", 4:"情绪输出", 5:"问题反馈"}
# 固定表头
FIXED_HEADERS = ["时间（TI）", "玩家ID（PID）", "玩家消息（PM）"]

# ========= 样式与尺寸 =========
# 字体（若本机没有“微雅软黑”，可改为“微软雅黑”等已安装字体）
FONT_NAME = "微雅软黑"
HEADER_FONT_SIZE = 12
BODY_FONT_SIZE   = 11

# 列宽（按表头名称配置）
COLUMN_WIDTHS = {
    "时间（TI）": 21,
    "玩家ID（PID）": 30,
    "玩家消息（PM）": 95,
}
# 表头行高（单位：磅），None 表示不改
HEADER_ROW_HEIGHT = 24
# 冻结首行（True -> 冻结到 A2）
FREEZE_HEADER = True
# 表头加粗、居中、底色（可选）
HEADER_BOLD = True
HEADER_FILL = "DDDDDD"  # 灰色；设为 None 不上底色

# 正文单元格边框（框线）
BORDER_ENABLE = True
BORDER_INCLUDE_HEADER = False   # True=表头也画框
BORDER_STYLE = "thin"           # thin / hair / medium / thick ...
# =====================================================

def log(msg): print(msg, flush=True)

def read_input_json_or_jsonl(path: Path, encoding="utf-8-sig") -> pd.DataFrame:
    text = path.read_text(encoding=encoding).strip()
    def to_df(records):
        df = pd.DataFrame(records)
        df = df.rename(columns={"时间":"TI","玩家ID":"PID","玩家消息":"PM","分类":"CA"})
        for col in ["TI","PID","PM","CA"]:
            if col not in df.columns:
                df[col] = "" if col!="CA" else pd.NA
        df = df[["TI","PID","PM","CA"]].copy()
        df["CA"] = pd.to_numeric(df["CA"], errors="coerce")
        return df
    try:
        obj = json.loads(text)
        if isinstance(obj, dict) and isinstance(obj.get("data"), list): return to_df(obj["data"])
        if isinstance(obj, list): return to_df(obj)
    except Exception: pass
    records=[]
    for i,line in enumerate(text.splitlines(),1):
        s=line.strip()
        if not s: continue
        try: records.append(json.loads(s))
        except Exception as e: print(f"[WARN] 第{i}行不是合法JSON：{e}", file=sys.stderr)
    if not records: raise ValueError("无法解析输入：请确认为有效 JSON/JSONL。")
    return to_df(records)

def ensure_workbook_and_sheets(xlsx: Path):
    if not xlsx.exists():
        wb=Workbook()
        first=True
        for ca in range(1,6):
            name=CA_TO_SHEET[ca]
            if first: ws=wb.active; ws.title=name; first=False
            else: ws=wb.create_sheet(title=name)
            ws.append(FIXED_HEADERS)
        wb.save(xlsx); return
    wb=load_workbook(xlsx)
    existing=set(wb.sheetnames)
    for name in CA_TO_SHEET.values():
        if name not in existing:
            ws=wb.create_sheet(title=name); ws.append(FIXED_HEADERS)
        else:
            ws=wb[name]
            if ws.max_row==0: ws.append(FIXED_HEADERS)
            else:
                header=[ws.cell(1,i+1).value for i in range(len(FIXED_HEADERS))]
                if header!=FIXED_HEADERS:
                    for i,h in enumerate(FIXED_HEADERS,1): ws.cell(1,i).value=h
    wb.save(xlsx)

def to_datetime_series(s: pd.Series, fmt: str|None):
    return pd.to_datetime(s, format=fmt, errors="coerce") if fmt else pd.to_datetime(s, errors="coerce")

def sort_by_time_cn(df: pd.DataFrame, time_format: str|None) -> pd.DataFrame:
    df["_ti"]=to_datetime_series(df["时间（TI）"], time_format)
    return df.sort_values("_ti").drop(columns=["_ti"])

def apply_sheet_format(xlsx_path: Path, sheet_name: str):
    """
    写入完成后，统一设置：
    - 列宽（按 FIXED_HEADERS 顺序映射 COLUMN_WIDTHS）
    - 表头行高
    - 冻结首行
    - 表头加粗/居中/底色（可选）
    - 正文单元格边框（框线）
    - 全表字体：表头 12、正文 11（均为 FONT_NAME）
    """
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    # 列宽：按表头名找对应宽度
    for idx, header in enumerate(FIXED_HEADERS, start=1):
        width = COLUMN_WIDTHS.get(header)
        if width:
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width

    # 表头行高
    if HEADER_ROW_HEIGHT:
        ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # 冻结首行
    if FREEZE_HEADER:
        ws.freeze_panes = "A2"

    # 表头样式（字体+对齐+底色）
    if HEADER_BOLD or HEADER_FILL:
        header_font = Font(name=FONT_NAME, size=HEADER_FONT_SIZE, bold=HEADER_BOLD)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill = PatternFill("solid", fgColor=HEADER_FILL) if HEADER_FILL else None
        for i in range(1, len(FIXED_HEADERS)+1):
            cell = ws.cell(row=1, column=i)
            cell.font = header_font
            cell.alignment = align
            if fill: cell.fill = fill

    # 正文字体（从第2行开始）
    body_font = Font(name=FONT_NAME, size=BODY_FONT_SIZE)
    max_row = ws.max_row
    end_col = len(FIXED_HEADERS)
    if max_row >= 2:
        for r in range(2, max_row+1):
            for c in range(1, end_col+1):
                cell = ws.cell(row=r, column=c)
                # 若之前设置过 alignment（自动换行/居中），保持；仅改字体
                cell.font = body_font

    # —— 正文（以及可选表头）单元格边框（框线）——
    if BORDER_ENABLE:
        side = Side(style=BORDER_STYLE, color="000000")  # 黑色边框
        border = Border(left=side, right=side, top=side, bottom=side)
        start_row = 1 if BORDER_INCLUDE_HEADER else 2
        if max_row >= start_row:
            for r in range(start_row, max_row + 1):
                for c in range(1, end_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    if r >= 2:
                        # 正文自动换行、垂直居中
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(xlsx_path)

def append_to_sheet(xlsx: Path, sheet_name: str, inc_df: pd.DataFrame,
                    dedup: bool, time_format: str|None):
    try:
        old = pd.read_excel(xlsx, sheet_name=sheet_name, dtype=str)
    except Exception:
        old = pd.DataFrame(columns=FIXED_HEADERS)

    merged = pd.concat([old, inc_df], ignore_index=True)
    if dedup:
        merged = merged.drop_duplicates(subset=["时间（TI）","玩家ID（PID）","玩家消息（PM）"], keep="first")
    merged = sort_by_time_cn(merged, time_format)

    # 覆盖 sheet
    wb = load_workbook(xlsx)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    wb.create_sheet(sheet_name)
    wb.save(xlsx)

    with pd.ExcelWriter(xlsx, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        merged = merged.reindex(columns=FIXED_HEADERS)
        merged.to_excel(writer, sheet_name=sheet_name, index=False)

    # 写完再设置格式（列宽/行高/冻结/样式/字体/边框）
    apply_sheet_format(xlsx, sheet_name)

def main():
    in_path = Path(INPUT_FILE)
    xlsx_path = Path(EXCEL_FILE)
    if not in_path.exists():
        print(f"[ERROR] 输入不存在：{in_path}", file=sys.stderr); sys.exit(1)

    if DO_BACKUP and xlsx_path.exists():
        bak = xlsx_path.with_suffix(xlsx_path.suffix + ".bak")
        shutil.copy2(xlsx_path, bak); log(f"[备份] 已创建：{bak}")

    new_df = read_input_json_or_jsonl(in_path, encoding=ENCODING)
    new_df = new_df.rename(columns={"TI":"时间（TI）","PID":"玩家ID（PID）","PM":"玩家消息（PM）"})
    new_df = new_df[["时间（TI）","玩家ID（PID）","玩家消息（PM）","CA"]].copy()

    ensure_workbook_and_sheets(xlsx_path)

    dedup = (not ALLOW_DUPLICATE)
    total = 0
    for ca, sheet in CA_TO_SHEET.items():
        inc = new_df[new_df["CA"]==ca][["时间（TI）","玩家ID（PID）","玩家消息（PM）"]].copy()
        if inc.empty: continue
        append_to_sheet(xlsx_path, sheet, inc, dedup, TIME_FORMAT)
        total += len(inc); log(f"[OK] 追加 {len(inc)} 到《{sheet}》")

    log(f"[完成] 共追加 {total} 条。Excel：{xlsx_path.resolve()}")

if __name__ == "__main__":
    main()
