# -*- coding: utf-8 -*-
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# ===========================================================
# =============== 工具函数 ===================================
# ===========================================================
def norm(s):
    """清除空字符"""
    return str(s).replace("\u200b", "").replace("\ufeff", "").strip() if s else ""

def strip_emoji(s: str) -> str:
    """去除 emoji 与无关符号"""
    if not s:
        return ""
    s = str(s)
    emoji_pattern = re.compile(
        "["
        "\U0001F300-\U0001F5FF"
        "\U0001F600-\U0001F64F"
        "\U0001F680-\U0001F6FF"
        "\U0001F700-\U0001F77F"
        "\U0001F780-\U0001F7FF"
        "\U0001F800-\U0001F8FF"
        "\U0001F900-\U0001F9FF"
        "\U0001FA00-\U0001FA6F"
        "\U0001FA70-\U0001FAFF"
        "\u200d"
        "\ufe0f"
        "]+", flags=re.UNICODE)
    s = emoji_pattern.sub("", s)
    s = re.sub(r"[\u3000\u200B\u200C\u200D]", "", s)
    return s.strip()

def normalize_time_str(t_raw: str) -> str:
    """多格式自动标准化为 yyyy-MM-dd HH:MM:SS"""
    if not t_raw:
        return ""
    t_raw = str(t_raw).strip().replace("/", "-")
    # 自动补秒
    parts = t_raw.split(" ")
    if len(parts) == 2:
        date_part, time_part = parts
        segs = time_part.split(":")
        if len(segs) == 2:
            time_part += ":00"
        elif len(segs) == 1:
            time_part += ":00:00"
        h, m, s = segs[0].zfill(2), segs[1].zfill(2), segs[2].zfill(2)
        t_raw = f"{date_part} {h}:{m}:{s}"
    try:
        dt = datetime.strptime(t_raw, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return t_raw

def parse_txt(txt_path):
    """解析TXT聊天记录"""
    print(f"🧾 正在读取 TXT：{txt_path}")
    pat = re.compile(r"(\d{4}[-/]\d{2}[-/]\d{2}\s+\d{1,2}:\d{2}(?::\d{2})?)\s+(.+)")
    recs = []
    with open(txt_path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.read().splitlines()
    i = 0
    while i < len(lines):
        m = pat.match(lines[i].strip())
        if m:
            t_raw, pid = m.groups()
            t_std = normalize_time_str(t_raw)
            msg = lines[i + 1].strip() if i + 1 < len(lines) else ""
            recs.append({
                "发言时间": t_std,
                "玩家ID": norm(pid),
                "玩家消息": strip_emoji(norm(msg))
            })
        i += 1
    df = pd.DataFrame(recs)
    print(f"📄 TXT 解析完成，共 {len(df)} 条（时间已标准化）")
    return df

def find_top_left_cell(ws, row, col):
    """在合并单元格下返回左上角单元格"""
    for rng in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(rng))
        if min_r <= row <= max_r and min_c <= col <= max_c:
            return ws.cell(row=min_r, column=min_c)
    return ws.cell(row=row, column=col)

# ===========================================================
# =============== 主逻辑 ====================================
# ===========================================================
def fill_empty_cells(excel_path, txt_path, output_path):
    df_txt = parse_txt(txt_path)
    wb = load_workbook(excel_path)
    print(f"📘 Excel 加载成功，检测到 {len(wb.sheetnames)} 个 sheet。")

    filled_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"🧩 开始处理 sheet：《{sheet_name}》 ...")

        data = ws.values
        try:
            cols = next(data)
        except StopIteration:
            continue

        df = pd.DataFrame(data, columns=cols)
        for c in ["话题簇", "发言时间", "玩家ID", "玩家消息"]:
            if c not in df.columns:
                df[c] = ""

        for i, row in df.iterrows():
            t0 = normalize_time_str(norm(row["发言时间"]))
            pid0 = norm(row["玩家ID"])
            msg0 = strip_emoji(norm(row["玩家消息"]))

            empty_cols = [c for c in ["发言时间", "玩家ID", "玩家消息"] if not norm(row[c])]
            if not empty_cols or len(empty_cols) == 3:
                continue

            ref = None

            # === 情况1：时间存在 ===
            if t0:
                match = df_txt[df_txt["发言时间"] == t0]
                if pid0:
                    match = match[match["玩家ID"] == pid0]
                elif msg0:
                    match = match[match["玩家消息"] == msg0]
                if not match.empty:
                    ref = match.iloc[0]

            # === 情况2：时间缺失，尝试用ID+消息定位 ===
            if not t0 and pid0 and msg0:
                match = df_txt[(df_txt["玩家ID"] == pid0) & (df_txt["玩家消息"] == msg0)]
                if not match.empty:
                    ref = match.iloc[0]
            # 仅ID匹配
            elif not t0 and pid0 and not msg0:
                match = df_txt[df_txt["玩家ID"] == pid0]
                if not match.empty:
                    ref = match.iloc[0]
            # 仅消息匹配（模糊匹配）
            elif not t0 and msg0 and not pid0:
                msg_core = msg0[:6]
                match = df_txt[df_txt["玩家消息"].str.contains(re.escape(msg_core), na=False)]
                if not match.empty:
                    ref = match.iloc[0]

            if ref is not None:
                for col_name in empty_cols:
                    new_val = ref[col_name]
                    if new_val:
                        col_idx = list(df.columns).index(col_name) + 1
                        cell = find_top_left_cell(ws, i + 2, col_idx)
                        cell.value = new_val
                        filled_count += 1

        print(f"✅ Sheet 《{sheet_name}》 补齐完成。")

    wb.save(output_path)
    print(f"🎯 空值补齐完成，共填充 {filled_count} 个单元格。")
    print(f"📁 输出文件：{output_path}")

# ===========================================================
# =============== 执行入口 ==================================
# ===========================================================
if __name__ == "__main__":
    excel_path = r"E:\项目\玩家社群分析智能体\玩家发言分类（供研发侧）\12232群_空值补齐V2修正.xlsx"
    txt_path   = r"E:\项目\玩家社群分析智能体\玩家发言分类（供研发侧）\1224《欢迎来到地球》测试2群.txt"
    output_path= excel_path.replace(".xlsx", "_空值补齐V2.xlsx")
    fill_empty_cells(excel_path, txt_path, output_path)
