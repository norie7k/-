# -*- coding: utf-8 -*-
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import PatternFill

# ===========================================================
# =============== å·¥å…·å‡½æ•° ===================================
# ===========================================================

def norm(s):
    """åŸºç¡€æ¸…æ´—"""
    return str(s).replace("\u200b", "").replace("\ufeff", "").strip() if s else ""

def strip_emoji(s: str) -> str:
    """å»é™¤ emojiã€å˜ä½“é€‰æ‹©ç¬¦ã€ç‰¹æ®Šç¬¦å·"""
    if not s:
        return ""
    s = str(s)
    emoji_pattern = re.compile(
        "["
        "\U0001F300-\U0001F5FF"
        "\U0001F600-\U0001F64F"
        "\U0001F680-\U0001F6FF"
        "\U0001F700-\U0001F77F"
        "\U0001F780-\U0001F7FF"
        "\U0001F800-\U0001F8FF"
        "\U0001F900-\U0001F9FF"
        "\U0001FA00-\U0001FA6F"
        "\U0001FA70-\U0001FAFF"
        "\u200d"
        "\u2640-\u2642"
        "\u2600-\u2B55"
        "\ufe0f"
        "]+", flags=re.UNICODE)
    s = emoji_pattern.sub("", s)
    s = re.sub(r"[\u3000\u200B\u200C\u200D]", "", s)
    return s.strip()

def has_ji(s):
    """æ£€æµ‹æ˜¯å¦å«â€œæâ€"""
    return "æ" in str(s) if s else False

def normalize_time_str(t_raw: str) -> str:
    """ç»Ÿä¸€æ—¶é—´æ ¼å¼ä¸º yyyy-MM-dd HH:MM:SS"""
    t_raw = norm(t_raw)
    m = re.match(r"(\d{4}-\d{2}-\d{2})\s+(\d{1,2}):(\d{2}):(\d{2})$", t_raw)
    if not m:
        raise ValueError(f"æ— æ³•è§£ææ—¶é—´: {t_raw}")
    date, hh, mm, ss = m.groups()
    hh = hh.zfill(2)
    return f"{date} {hh}:{mm}:{ss}"

def parse_txt(path: str) -> pd.DataFrame:
    """è§£æ TXT ä¸º DataFrame å¹¶æ¸…æ´— emojiã€æ ‡å‡†åŒ–æ—¶é—´"""
    pat_head = re.compile(r"(\d{4}-\d{2}-\d{2}\s+\d{1,2}:\d{2}:\d{2})\s+(.+)")
    recs = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.read().splitlines()
    n = len(lines)
    i = 0
    while i < n:
        line = lines[i].strip()
        m = pat_head.match(line)
        if m:
            t_raw, pid = m.groups()
            try:
                t_fmt = normalize_time_str(t_raw)
            except Exception:
                i += 1
                continue
            msg = lines[i + 1].strip() if (i + 1) < n else ""
            recs.append({
                "å‘è¨€æ—¶é—´": t_fmt,
                "ç©å®¶ID": norm(pid),
                "ç©å®¶æ¶ˆæ¯": strip_emoji(norm(msg))
            })
        i += 1
    df = pd.DataFrame(recs, columns=["å‘è¨€æ—¶é—´","ç©å®¶ID","ç©å®¶æ¶ˆæ¯"])
    print(f"ğŸ“„ TXT è§£æå®Œæˆï¼š{len(df)} æ¡ï¼Œæ—¶é—´ç»Ÿä¸€ä¸º yyyy-MM-dd HH:MM:SS")
    return df

def find_top_left_cell(ws, row, col):
    """åˆå¹¶å•å…ƒæ ¼å®‰å…¨å†™å…¥"""
    for rng in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(rng))
        if min_r <= row <= max_r and min_c <= col <= max_c:
            return ws.cell(row=min_r, column=min_c)
    return ws.cell(row=row, column=col)

# ===========================================================
# =============== ä¸»é€»è¾‘ ====================================
# ===========================================================

def fix_extreme_by_rule(excel_path, txt_path, output_path, log_path):
    wb = load_workbook(excel_path)
    df_txt = parse_txt(txt_path)
    if not isinstance(df_txt, pd.DataFrame) or df_txt.empty:
        raise RuntimeError("âŒ TXT è§£æå¤±è´¥æˆ–ä¸ºç©ºï¼Œè¯·æ£€æŸ¥ txt æ–‡ä»¶æ ¼å¼ã€‚")

    fill_yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    logs = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        data = ws.values
        try:
            cols = list(next(data))  # ç¬¬ä¸€è¡Œè¡¨å¤´
        except StopIteration:
            continue

        df = pd.DataFrame(data, columns=cols)
        for c in ["è¯é¢˜ç°‡","å‘è¨€æ—¶é—´","ç©å®¶ID","ç©å®¶æ¶ˆæ¯"]:
            if c not in df.columns:
                df[c] = ""

        print(f"ğŸ§© å¼€å§‹å¤„ç† sheetï¼šã€Š{sheet}ã€‹ ...")

        # -------- â‘  æ¸…ç†è¯é¢˜ç°‡ä¸­çš„â€œæâ€ã€â€œææµ·å¬é›·â€ --------
        for idx, val in enumerate(df["è¯é¢˜ç°‡"]):
            if isinstance(val, str) and ("æ" in val or "ææµ·å¬é›·" in val):
                new_val = val.replace("ææµ·å¬é›·", "").replace("æ", "").strip()
                if new_val != val:
                    c_idx = cols.index("è¯é¢˜ç°‡") + 1
                    cell = find_top_left_cell(ws, idx + 2, c_idx)
                    cell.value = new_val
                    cell.fill = fill_yellow
                    logs.append({
                        "sheet": sheet,
                        "è¡Œå·": idx + 2,
                        "å­—æ®µ": "è¯é¢˜ç°‡",
                        "åŸå€¼": val,
                        "ä¿®æ­£å€¼": new_val,
                        "ä¾æ®æ—¶é—´": ""
                    })

        # -------- â‘¡ å«æä¿®æ­£é€»è¾‘ï¼ˆå¢å¼ºç‰ˆï¼šå‘è¨€æ—¶é—´+ID+æ¶ˆæ¯ä¿®æ­£ï¼‰ --------
        fix_cols = ["å‘è¨€æ—¶é—´", "ç©å®¶ID", "ç©å®¶æ¶ˆæ¯"]  # è¯é¢˜ç°‡ä¸åœ¨ç¬¬äºŒè½®è¦†ç›–

        for i, row in df.iterrows():
            t_raw = norm(row.get("å‘è¨€æ—¶é—´"))
            pid   = strip_emoji(norm(row.get("ç©å®¶ID")))
            msg   = strip_emoji(norm(row.get("ç©å®¶æ¶ˆæ¯")))
            topic = norm(row.get("è¯é¢˜ç°‡"))

            # å››å­—æ®µæ˜¯å¦å«æ
            t_bad     = has_ji(t_raw)
            id_bad    = has_ji(pid)
            msg_bad   = has_ji(msg)
            topic_bad = has_ji(topic)

            if not (t_bad or id_bad or msg_bad or topic_bad):
                continue

            # å‘è¨€æ—¶é—´å«æ â†’ å…ˆæ¸…æ´—æˆå¯åŒ¹é… TXT çš„æ—¶é—´æ ¼å¼
            t_clean = t_raw.replace("ææµ·å¬é›·","").replace("æ","æé€Ÿèœ—ç‰›").strip()
            if not t_clean:
                continue

            # åŒ¹é… TXT åŸå§‹è®°å½•
            cand = df_txt[df_txt["å‘è¨€æ—¶é—´"] == t_clean]
            if cand.empty:
                continue
            ref = cand.iloc[0]

            # ä¿®æ­£â€œå‘è¨€æ—¶é—´ / ç©å®¶ID / ç©å®¶æ¶ˆæ¯â€
            for col in fix_cols:
                old_val = norm(row.get(col))
                if has_ji(old_val):  # åªä¿®å«æçš„
                    new_val = norm(ref.get(col))
                    if new_val and new_val != old_val:
                        c_idx = cols.index(col) + 1
                        cell = find_top_left_cell(ws, i + 2, c_idx)
                        cell.value = new_val
                        cell.fill = fill_yellow
                        logs.append({
                            "sheet": sheet,
                            "è¡Œå·": i + 2,
                            "å­—æ®µ": col,
                            "åŸå€¼": old_val,
                            "ä¿®æ­£å€¼": new_val,
                            "ä¾æ®æ—¶é—´": t_clean
                        })

        print(f"âœ… Sheet ã€Š{sheet}ã€‹ å¤„ç†å®Œæˆã€‚")

    wb.save(output_path)
    print(f"ğŸ’¾ ä¿®æ­£å®Œæˆï¼š{output_path}")

    if logs:
        pd.DataFrame(logs).to_excel(log_path, index=False)
        print(f"ğŸ“ ä¿®æ­£æ—¥å¿—ç”Ÿæˆï¼š{log_path}")
    else:
        print("âšª æœªå‘ç°éœ€è¦ä¿®æ­£çš„å­—æ®µã€‚")

# ===========================================================
# =============== å…¥å£ =======================================
# ===========================================================
if __name__ == "__main__":
    excel_path = r"E:\é¡¹ç›®\ç©å®¶ç¤¾ç¾¤åˆ†ææ™ºèƒ½ä½“\ç©å®¶å‘è¨€åˆ†ç±»ï¼ˆä¾›ç ”å‘ä¾§ï¼‰\12232ç¾¤_ç©ºå€¼è¡¥é½V2.xlsx"
    txt_path   = r"E:\é¡¹ç›®\ç©å®¶ç¤¾ç¾¤åˆ†ææ™ºèƒ½ä½“\ç©å®¶å‘è¨€åˆ†ç±»ï¼ˆä¾›ç ”å‘ä¾§ï¼‰\1224ã€Šæ¬¢è¿æ¥åˆ°åœ°çƒã€‹æµ‹è¯•2ç¾¤.txt"
    output_path= excel_path.replace(".xlsx", "ä¿®æ­£.xlsx")
    log_path   = excel_path.replace(".xlsx", "_å«æä¸‰æ€ä¿®æ­£.xlsx")
    fix_extreme_by_rule(excel_path, txt_path, output_path, log_path)
    print("ğŸ¯ å…¨æµç¨‹æ‰§è¡Œå®Œæ¯•ï¼")
