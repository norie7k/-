# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from copy import copy

# ======================== é…ç½® =========================
SHEET_NAMES = ["ä½“éªŒåé¦ˆ", "ç–‘æƒ‘è¯¢é—®", "å»ºè®®çµæ„Ÿ", "æƒ…ç»ªè¾“å‡º", "é—®é¢˜åé¦ˆ"]
HEADERS     = ["è¯é¢˜ç°‡", "å‘è¨€æ—¶é—´", "ç©å®¶ID", "ç©å®¶æ¶ˆæ¯"]

# æ ·å¼é…ç½®
FONT_NAME         = "å¾®é›…è½¯é»‘"      # è‹¥æœ¬æœºæ²¡æœ‰ï¼Œå¯æ”¹ "å¾®è½¯é›…é»‘"
HEADER_FONT_SIZE  = 16
BODY_FONT_SIZE    = 11
HEADER_FILL       = "FFDDEBF7"     # æµ…è“åº•
HEADER_ROW_HEIGHT = 24
COL_WIDTHS        = [16, 21, 30, 95]   # å¯¹åº”å„åˆ—å®½åº¦

# ======================== å·¥å…·å‡½æ•° =========================
def copy_cell_style(src_cell, dest_cell):
    """å¤åˆ¶å•å…ƒæ ¼æ ·å¼"""
    if src_cell and src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = copy(src_cell.number_format)
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)

def set_sheet_format(ws):
    """åº”ç”¨ç»Ÿä¸€æ ·å¼ï¼šå†»ç»“é¦–è¡Œã€åˆ—å®½ã€è¡¨å¤´æ ·å¼"""
    ws.freeze_panes = "A2"
    for i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    # è¡¨å¤´æ ·å¼
    for c, text in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c)
        cell.value = text
        cell.font = Font(name=FONT_NAME, bold=True, size=HEADER_FONT_SIZE)
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

def extract_topic_blocks(ws):
    """
    ä»ä¸€ä¸ª sheet ä¸­æå–æ‰€æœ‰ã€Œè¯é¢˜ç°‡å—ã€ä¿¡æ¯ï¼š
    æ¯ä¸ªå—åŒ…å«ï¼š
        {
            'topic': è¯é¢˜ç°‡åç§°,
            'rows': [[æ¯åˆ—çš„å€¼], ...],
            'merged': [(min_col, min_row, max_col, max_row)]
        }
    """
    topic_blocks = []
    topic_col = 1  # è¯é¢˜ç°‡åˆ—ä¸ºç¬¬ä¸€åˆ—
    merges = list(ws.merged_cells.ranges)
    nrows, ncols = ws.max_row, ws.max_column

    merged_row_set = set()
    # æ”¶é›†æ‰€æœ‰åˆå¹¶è¡Œçš„èŒƒå›´
    for rng in merges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        for r in range(min_row, max_row + 1):
            merged_row_set.add(r)
        # ä»…å¤„ç†è¯é¢˜ç°‡åˆ—çš„åˆå¹¶åŒºåŸŸ
        if min_col == max_col == topic_col:
            topic = ws.cell(row=min_row, column=topic_col).value
            block_rows = []
            for r in range(min_row, max_row + 1):
                block_rows.append([ws.cell(row=r, column=c).value for c in range(1, ncols + 1)])
            topic_blocks.append({
                "topic": topic or "",
                "rows": block_rows,
                "merged": [(min_col, min_row, max_col, max_row)]
            })

    # æ‰¾å‡ºæœªåˆå¹¶çš„è¡Œï¼ˆå•ç‹¬è¯é¢˜ç°‡ï¼‰
    for r in range(2, nrows + 1):
        if r not in merged_row_set:
            topic = ws.cell(row=r, column=topic_col).value or ""
            row_data = [[ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]]
            topic_blocks.append({
                "topic": topic,
                "rows": row_data,
                "merged": []  # æ— åˆå¹¶ç»“æ„
            })

    return topic_blocks


def copy_block_to(ws_dest, ws_src, block, start_row):
    """å°†æ•´ä¸ªè¯é¢˜ç°‡å—å¤åˆ¶åˆ°ç›®æ ‡ sheetï¼ˆä¿ç•™æ ·å¼+åˆå¹¶ç»“æ„ï¼‰"""
    base_row = start_row
    ncols = ws_src.max_column

    # ğŸ”¹ å¦‚æœè¯¥å—æ— åˆå¹¶ç»“æ„ï¼ˆå•ç‹¬è¡Œï¼‰
    if not block["merged"]:
        for i, row_values in enumerate(block["rows"], start=0):
            for j, val in enumerate(row_values, start=1):
                src = ws_src.cell(row=2, column=j)  # ä»ç¬¬2è¡Œå–æ ·å¼æ¨¡æ¿
                dest = ws_dest.cell(row=base_row + i, column=j, value=val)
                copy_cell_style(src, dest)
        return

    # ğŸ”¹ æœ‰åˆå¹¶ç»“æ„çš„å—
    min_col, min_row, max_col, max_row = block["merged"][0]
    for i, row_values in enumerate(block["rows"], start=0):
        for j, val in enumerate(row_values, start=1):
            src = ws_src.cell(row=min_row + i, column=j)
            dest = ws_dest.cell(row=base_row + i, column=j, value=val)
            copy_cell_style(src, dest)

    # é‡å»ºåˆå¹¶ç»“æ„
    for (min_col, _, max_col, max_row) in block["merged"]:
        height = max_row - min_row
        ws_dest.merge_cells(
            start_row=base_row,
            end_row=base_row + height,
            start_column=min_col,
            end_column=max_col
        )

# ======================== ä¸»é€»è¾‘ =========================
def merge_excels_keep_blocks_sorted(excel_files, output_path):
    print(f"ğŸ§© åˆå¹¶ {len(excel_files)} ä¸ª Excelï¼ŒæŒ‰è¯é¢˜ç°‡å—æ’åºï¼Œä¿ç•™åŸåˆå¹¶ç»“æ„...")

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in SHEET_NAMES:
        print(f"\nğŸ”§ å¤„ç† Sheetï¼šã€Š{sheet_name}ã€‹ ...")
        all_blocks = []
        template_ws = None

        # æ”¶é›†æ‰€æœ‰æ–‡ä»¶ä¸­è¯¥ Sheet çš„è¯é¢˜ç°‡å—
        for path in excel_files:
            wb = openpyxl.load_workbook(path)
            if sheet_name not in wb.sheetnames:
                print(f"âš ï¸ {path} ç¼ºå°‘ Sheetï¼šã€Š{sheet_name}ã€‹ ï¼Œè·³è¿‡ã€‚")
                continue
            ws = wb[sheet_name]
            if template_ws is None:
                template_ws = ws
            blocks = extract_topic_blocks(ws)
            all_blocks.extend(blocks)

        # æŒ‰è¯é¢˜ç°‡åç§°å‡åºæ’åºï¼ˆç©ºå€¼æ’æœ€åï¼‰
        all_blocks.sort(key=lambda b: (b["topic"] == "", b["topic"] or ""))

        # å†™å…¥æ–°å·¥ä½œç°¿
        ws_new = wb_out.create_sheet(title=sheet_name)
        set_sheet_format(ws_new)

        current_row = 2
        for block in all_blocks:
            copy_block_to(ws_new, template_ws, block, current_row)
            current_row += len(block["rows"])

        print(f"âœ… Sheet ã€Š{sheet_name}ã€‹ å®Œæˆï¼šå…± {len(all_blocks)} ä¸ªè¯é¢˜ç°‡å—ã€‚")

    wb_out.save(output_path)
    print(f"\nğŸ¯ æ‰€æœ‰ Sheet åˆå¹¶å®Œæˆï¼Œä¿ç•™åˆå¹¶æ ¼å¼è¾“å‡ºï¼š{output_path}")

# ======================== æ‰§è¡Œå…¥å£ =========================


# ======================== æ‰§è¡Œå…¥å£ =========================
if __name__ == "__main__":
    excel_files = [
        r"E:\é¡¹ç›®\ç©å®¶ç¤¾ç¾¤åˆ†ææ™ºèƒ½ä½“\ç©å®¶å‘è¨€åˆ†ç±»ï¼ˆä¾›ç ”å‘ä¾§ï¼‰\01052ç¾¤_ç©ºå€¼è¡¥é½V2ä¿®æ­£.xlsx",
        r"E:\é¡¹ç›®\ç©å®¶ç¤¾ç¾¤åˆ†ææ™ºèƒ½ä½“\ç©å®¶å‘è¨€åˆ†ç±»ï¼ˆä¾›ç ”å‘ä¾§ï¼‰\ã€åœ°çƒã€‘_ã€Œæš–å†¬æµ‹è¯•ã€_ç©å®¶å‘è¨€åˆ†ç±»ã€Œ2ç¾¤ã€_251231~260104ï¼ˆæŒç»­æ›´æ–°ï¼‰.xlsx"
        
    ]
    output_path = r"E:\é¡¹ç›®\ç©å®¶ç¤¾ç¾¤åˆ†ææ™ºèƒ½ä½“\ç©å®¶å‘è¨€åˆ†ç±»ï¼ˆä¾›ç ”å‘ä¾§ï¼‰\ã€åœ°çƒã€‘_ã€Œæš–å†¬æµ‹è¯•ã€_ç©å®¶å‘è¨€åˆ†ç±»ã€Œ2ç¾¤ã€_251231~260105ï¼ˆæŒç»­æ›´æ–°ï¼‰.xlsx"
    merge_excels_keep_blocks_sorted(excel_files, output_path)



 


  