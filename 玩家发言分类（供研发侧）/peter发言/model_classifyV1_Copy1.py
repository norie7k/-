
from __future__ import annotations
import json, time, typing as T
import pandas as pd
import requests
from pathlib import Path
from typing import List, Dict, Any
import re, json, unicodedata
import json
# --- openpyxl 样式/工具 ---
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle

import re

################模型调用，出结果###################

def load_system_prompt(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Prompt file not found: {path}")
    return path.read_text(encoding="utf-8")

def build_user_prompt_single_speaker_behavior(batch_lines: list[str]) -> str:
    jsonl_block = "\n".join([line.strip() for line in batch_lines if str(line).strip()])
    return (
        "请严格按照【系统提示词】执行单人社群表达路径行为分析。\n"
        "注意：禁止复述/粘贴输入JSONL，禁止只输出原始JSON行；必须输出系统要求的五段结构化结果。\n\n"
        "【输入JSONL】\n"
        f"{jsonl_block}"
    )


def call_ark_chat_completions(
    api_url: str,
    api_key: str,
    model: str,
    system_prompt: str,
    user_prompt: str,
    temperature: float = 0.3,
    max_tokens: int = 32700,
    timeout: int = 600,
    retries: int = 2,
) -> str:
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    last_err = None
    for attempt in range(retries + 1):
        try:
            resp = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text}")
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            last_err = e
            time.sleep(1.2 * (attempt + 1))
    raise RuntimeError(f"Ark API 调用失败: {last_err}")

def extract_valid_json_lines(text: str) -> T.List[str]:
    """
    把模型输出里的纯 JSON 行提取出来（鲁棒处理）：
    - 逐行判断：以 { 开头 且 以 } 结尾，则认为是一个 JSON 对象行
    - 也能容忍前后多余空行或解释文字（会被忽略）
    """
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("{") and s.endswith("}"):
            lines.append(s)
    return lines






