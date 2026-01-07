from __future__ import annotations
import json, time, typing as T
import pandas as pd
import requests
import math
from collections import defaultdict
from pathlib import Path
from typing import List, Dict, Any, Optional
import re, unicodedata
from datetime import datetime
from json import JSONDecodeError

# 进度打印（parse_model2_output_to_json_list 用到）
from tqdm import tqdm

# --- openpyxl 样式/工具 ---
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle

# --- docx 样式/工具 ---
from docx.oxml import OxmlElement
from docx import Document
from docx.shared import Pt
from docx.oxml.ns import qn


################ 模型调用，出结果 ###################

def load_system_prompt(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Prompt file not found: {path}")
    return path.read_text(encoding="utf-8")


def build_user_prompt_filter(json_lines: T.List[str]) -> str:
    # 模型#1：筛掉非游戏相关，只输出相关 JSON 行（原样）
    return (
        "以下是若干玩家/客服/研发的发言记录，请根据系统提示中规则，"
        "判断哪些是【与游戏内容相关】的发言，保留这些 JSON 行，不相关的忽略。"
        "请仅输出【相关发言的原始 JSON 行】，严格保持格式不变。\n\n"
        "【输入】：\n" + "\n".join(json_lines)
    )


def build_user_prompt_clsuter(jsonl_block: str) -> str:
    return (
        "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请先完整阅读全部输入，然后按系统提示中的话题簇规则进行划分。\n"
        "【输出要求】只输出若干 JSON 对象，每个话题簇一个 JSON；"
        "禁止使用 ```json 或 ``` 等 Markdown 代码块，禁止输出解释文字。\n\n"
        "【输入】：\n" + jsonl_block
    )


def build_user_prompt_cluster_agg(jsonl_block: str) -> str:
    return (
        "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请先完整阅读全部输入，然后按系统提示中的话题簇规则进行划分。\n"
        "【输出要求】只输出若干 JSON 对象，每个话题簇一个 JSON；"
        "禁止使用 ```json 或 ``` 等 Markdown 代码块，禁止输出解释文字。\n\n"
        "【输入】：\n" + jsonl_block
    )


def build_user_prompt_version_agg(jsonl_block: str) -> str:
    return (
        "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请先完整阅读全部输入，然后按系统提示中的话题簇规则进行划分。\n"
        "【输出要求】只输出若干 JSON 对象，每个话题簇一个 JSON；"
        "禁止使用 ```json 或 ``` 等 Markdown 代码块，禁止输出解释文字。\n\n"
        "【输入】：\n" + jsonl_block
    )


def build_user_prompt_subcluster_opinion(
    topic_id: str,
    discussion_point: str,
    dialogs: List[Dict[str, Any]],
) -> str:
    """
    模型#4（观点分析）用户提示词构造：
    第一行是话题簇元信息，其余是原始发言 JSON 行。
    """
    lines: List[str] = []

    meta = {
        "话题簇ID": topic_id,
        "讨论点": discussion_point,
    }
    lines.append(json.dumps(meta, ensure_ascii=False))

    for row in dialogs:
        lines.append(json.dumps(row, ensure_ascii=False))

    jsonl_block = "\n".join(lines)

    return (
        "禁止使用 ```json 或 ``` 等 Markdown 代码块，禁止输出解释文字。\n\n"
        "【输入】本次所有数据（JSONL，第一行为话题簇信息，其余为发言）：\n"
        + jsonl_block +
        "\n\n仅根据以上内容，并遵循系统提示词中的规则和输出格式，"
        "直接输出 1 个 JSON 对象作为最终结果。"
    )

def build_user_prompt_version_opinion(
    discussion_point: str,
    json_lines: List[Any],   # 实际传的是 list[str]
) -> str:
    dp = (discussion_point or "").strip()
    jsonl_block = "\n".join(
        (line or "").strip() for line in json_lines if line
    )
    return (
        "你将收到一段玩家聊天原文（JSONL，每行一条）。本次只分析指定【讨论点】，其他内容忽略。\n\n"
        "【输出要求】只输出1个JSON对象，禁止解释文字，禁止Markdown代码块。\n\n"
        "【输入话题点】：\n" + dp + "\n\n"
        "【输入原文JSONL】\n" + jsonl_block
    )
def build_user_prompt_heat_trend(jsonl_block: str) -> str:
    """
    给模型#7 的用户提示词：
    - jsonl_block 可以是一行或多行 JSONL
    - 我们现在是一行一跑，所以实际就是一行
    """
    return (
        "以下是版本发言Top5中的一个【聚合话题簇】及其讨论点量化信息，"
        "格式为 JSONL（每行一个对象，本次只有一行）。\n\n"
        "请严格按照系统提示词《热度趋势智能体》的要求，只输出 1 个 JSON 对象："
        "包含字段【聚合话题簇】和【一句话总结】，"
        "禁止输出 Markdown 代码块，禁止解释文字。\n\n"
        "【输入】\n" + jsonl_block.strip()
    )

def call_ark_chat_completions(
    api_url: str,
    api_key: str,
    model: str,
    system_prompt: str,
    user_prompt: str,
    temperature: float = 0.3,
    max_tokens: int = 32700,
    timeout: int = 600,
    retries: int = 2,
) -> str:
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    last_err = None
    for attempt in range(retries + 1):
        try:
            resp = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text}")
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            last_err = e
            time.sleep(1.2 * (attempt + 1))
    raise RuntimeError(f"Ark API 调用失败: {last_err}")


def extract_valid_json_lines(text: str) -> T.List[str]:
    """
    把模型输出里的纯 JSON 行提取出来（鲁棒处理）：
    - 逐行判断：以 { 开头 且 以 } 结尾，则认为是一个 JSON 对象行
    - 也能容忍前后多余空行或解释文字（会被忽略）
    """
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("{") and s.endswith("}"):
            lines.append(s)
    return lines


def add_index_to_jsonl_lines(jsonl_lines: List[str]) -> List[str]:
    """
    给原始 jsonl 每条发言加上一个唯一行号字段 `_idx`，
    返回新的 List[str]，每个元素仍然是一行 JSON 字符串。
    """
    new_lines = []
    for idx, line in enumerate(jsonl_lines, start=1):
        line = line.strip()
        if not line:
            continue
        obj = json.loads(line)
        obj["_idx"] = idx  # 新增行号
        new_lines.append(json.dumps(obj, ensure_ascii=False))
    return new_lines


def count_output_filter_stats(output_filter: str):
    """
    统计模型#1 输出中的：
    - 总行数 total_lines（玩家+客服+研发）
    - 玩家发言行数 player_lines（简单用 发言人ID 里的关键词区分）
    """
    total_lines = 0
    player_lines = 0

    for line in output_filter.splitlines():
        line = line.strip()
        if not line:
            continue
        total_lines += 1

        try:
            obj = json.loads(line)
        except Exception:
            # 如果这一行不是合法 JSON，就只能算进 total_lines，无法细分
            continue

        speaker = (
            obj.get("玩家ID")
            or obj.get("发言人ID")
            or obj.get("角色ID")
            or ""
        )

        # 简单排除客服/GM/官方/运营/研发，其余视为玩家
        if any(key in str(speaker) for key in ["客服", "GM", "官方", "运营", "研发"]):
            continue

        player_lines += 1

    return total_lines, player_lines


def get_covered_indices_from_cluster_output(output_cluster: str):
    """
    从模型#2 的自然语言输出中，解析所有 “发言行号列表：[...数字...]” 里的数字，
    收集成一个 set 返回。

    适配类似格式：
        发言行号列表：[1, 2, 3]
        发言行号列表：[88, 90]
        发言行号列表：[124,125,126]
    """
    covered = set()

    # 正则匹配 “发言行号列表：[1, 2, 3]”
    pattern = r"发言行号列表[:：]\s*[\[\【]([0-9,\s，]+)[\]\】]"

    for m in re.finditer(pattern, output_cluster):
        nums_str = m.group(1)  # 里面是类似 "1, 2, 3" 或 "124,125,126"
        # 替换中文逗号，按逗号切分
        nums_str = nums_str.replace("，", ",")
        for part in nums_str.split(","):
            p = part.strip()
            if not p:
                continue
            try:
                covered.add(int(p))
            except ValueError:
                continue

    return covered


################################ 修补 bug — 话题簇划分解析 #########################

def _normalize_json_text(text: str) -> str:
    """
    对模型输出做一些小修复，以便 json.loads 正常解析：
    1）去掉整段末尾多余的逗号：{"a":1},
    2）去掉 属性/元素 后紧跟 } 或 ] 的非法逗号。
    """
    text = text.rstrip()

    # 1) 整个对象末尾多打一颗逗号：{"a":1},
    if text.endswith(","):
        text = text[:-1].rstrip()

    # 2) 属性/元素后直接跟 } 或 ] 的逗号
    text = re.sub(r",\s*(\n\s*[}\]])", r"\1", text)

    return text


def parse_model2_output_to_json_list(output_cluster: str, batch_idx: int = 0) -> List[Dict[str, Any]]:
    """
    把模型#2的原始输出解析成 list[dict]：
    - 支持一行一个 JSON：
        {"话题簇1": "...", "核心对象/机制": "..."}
    - 也支持多行漂亮 JSON：
        {
          "话题簇1": "...",
          "核心对象/机制": "..."
        }
    - 自动忽略解释性文字、markdown 列表 "- {...}"
    - 调用 _normalize_json_text 修复尾逗号等格式问题
    - 解析失败时打印完整对象原文，但不中断整批
    """
    objs: List[Dict[str, Any]] = []
    cur_lines: List[str] = []
    depth = 0  # 当前大括号嵌套层数

    lines = output_cluster.strip().splitlines()

    for idx, raw in enumerate(lines, start=1):
        line = (raw or "").rstrip()
        if not line.strip():
            continue

        # 去掉 markdown 列表前缀（常见格式："- {...}"）
        if line.lstrip().startswith("- "):
            line = line.lstrip()[2:].lstrip()

        # 当前还没进入任何 JSON 对象，并且这一行也看不到 "{"
        if depth == 0 and "{" not in line:
            # 大概率是解释性文字，比如 "以下是话题簇..."
            continue

        # 统计这一行的大括号数量
        open_cnt = line.count("{")
        close_cnt = line.count("}")

        if depth == 0:
            # 刚刚进入一个新的对象
            cur_lines = [line]
            depth = open_cnt - close_cnt

            # depth <= 0 说明这是“单行对象”：{"a":1} 或 {"a":1},
            if depth <= 0:
                text = "\n".join(cur_lines)
                text = _normalize_json_text(text)
                try:
                    obj = json.loads(text)
                except JSONDecodeError as e:
                    tqdm.write(f"[批次 {batch_idx}] ❌ JSON解析失败（行#{idx}附近）：{e}")
                    tqdm.write(f"[批次 {batch_idx}] 该对象原文：{text}")
                else:
                    # 统一 "话题簇1"/"话题簇2"... -> "话题簇"
                    for key in list(obj.keys()):
                        if key.startswith("话题簇") and key != "话题簇":
                            obj["话题簇"] = obj.pop(key)
                    objs.append(obj)
                cur_lines = []
                depth = 0
        else:
            # 已经在一个 JSON 对象内部（多行场景）
            cur_lines.append(line)
            depth += open_cnt - close_cnt

            # depth 回到 0，表示一个对象结束
            if depth <= 0:
                text = "\n".join(cur_lines)
                text = _normalize_json_text(text)

                try:
                    obj = json.loads(text)
                except JSONDecodeError as e:
                    tqdm.write(f"[批次 {batch_idx}] ❌ JSON解析失败（行#{idx}附近）：{e}")
                    tqdm.write(f"[批次 {batch_idx}] 该对象原文：{text}")
                else:
                    for key in list(obj.keys()):
                        if key.startswith("话题簇") and key != "话题簇":
                            obj["话题簇"] = obj.pop(key)
                    objs.append(obj)

                cur_lines = []
                depth = 0

    return objs


def fix_model3_line_extreme_axis(s: str) -> str:
    """
    专门修复模型#3输出中“极轴”相关的坏格式：
    1) "日期": "极轴": "2025-12-06"  =>  "日期": "2025-12-06"
    2) "时间轴": "极轴": "22:30:57-22:33:57"  =>  "时间轴": "22:30:57-22:33:57"
    3) "时间轴": "22:34:24-极轴": "22:38:33" => "时间轴": "22:34:24-22:38:33"
    """
    s = re.sub(r'"日期"\s*:\s*"极轴"\s*:\s*"([^"]+)"', r'"日期": "\1"', s)
    s = re.sub(r'"时间轴"\s*:\s*"极轴"\s*:\s*"([^"]+)"', r'"时间轴": "\1"', s)
    s = re.sub(r'"时间轴"\s*:\s*"([^"]*?)-极轴"\s*:\s*"([^"]+)"', r'"时间轴": "\1-\2"', s)
    return s


def parse_jsonl_text_safe(text: str, label: str = "模型#3聚合输出") -> List[Dict[str, Any]]:
    """
    尝试把 text 按行解析为 JSON 对象：
    - 每行独立尝试 json.loads
    - 解析前先对“极轴”等常见错误做一次正则修复
    - 解析失败时不会抛异常，而是打印出具体行号和原文，方便排查
    """
    objs: List[Dict[str, Any]] = []
    lines = text.strip().splitlines()

    for idx, raw in enumerate(lines, start=1):
        s = (raw or "").strip()
        if not s:
            continue

        # 跳过 ```json / ``` 等代码块标记
        if s.startswith("```"):
            continue

        # ⭐ 先修复“极轴”相关格式错误
        s_fixed = fix_model3_line_extreme_axis(s)

        try:
            obj = json.loads(s_fixed)
        except JSONDecodeError as e:
            print(f"[{label}] ❌ JSON解析失败：行#{idx} -> {e}")
            print(f"[{label}] 该行原文(修复后)：{s_fixed}")
            continue

        objs.append(obj)

    return objs


################################# 话题簇唯一 ID #################################

def infer_date_for_batch(
    cluster_json_list: List[Dict[str, Any]],
    batch_lines: List[str]
) -> str:
    """
    优先从话题簇标题中抽取日期（形如：xxxx（2025-12-02 16:30:01-17:42:41））
    若失败，则回退到原始发言中的 `发言日期` / `日期` 字段。
    """
    # 1）先从话题簇标题里找 YYYY-MM-DD
    for obj in cluster_json_list:
        title = str(
            obj.get("话题簇")
            or obj.get("聚合话题簇")
            or ""
        )
        m = re.search(r"(\d{4}-\d{2}-\d{2})", title)
        if m:
            return m.group(1)

    # 2）如果标题里没有，就从原始发言里拿
    for line in batch_lines:
        line = line.strip()
        if not line:
            continue
        try:
            msg = json.loads(line)
        except Exception:
            continue

        date_str = msg.get("发言日期") or msg.get("日期")
        if date_str:
            return date_str

    # 3）再不行就报错，让你意识到数据格式有问题
    raise ValueError("无法从话题簇标题或原始发言中推断出日期，请检查数据格式。")


def assign_global_cluster_ids(cluster_list: List[Dict[str, Any]], date_str: str, batch_id: int):
    """
    为每个话题簇生成全局唯一ID字段 `_cluster_id`
    格式：YYYY-MM-DD_BX_XX，如 2025-11-20_B2_03
    """
    for idx, cluster in enumerate(cluster_list, start=1):
        cluster["_cluster_id"] = f"{date_str}_{batch_id}_{idx:02d}"
    return cluster_list


################################# 聚合每天的话提簇分批输出 #################################

def aggregate_cluster_outputs(batch_outputs: List[str]) -> str:
    """
    把多批模型#2 输出聚合成一个 jsonl 字符串。
    """
    all_lines: List[str] = []

    for _, text in enumerate(batch_outputs, start=1):
        if not text:
            continue

        for line in text.strip().splitlines():
            line = line.strip()
            if not line:
                continue

            try:
                obj = json.loads(line)
            except Exception:
                # 如果有一行不是合法 JSON，就跳过
                continue

            clean_line = json.dumps(obj, ensure_ascii=False)
            all_lines.append(clean_line)

    # 聚合为一个大的 JSONL 字符串
    return "\n".join(all_lines)


################ 聚合修复“极轴” #############################


def _strip_fences(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"^```[a-zA-Z0-9]*\s*", "", s)
    s = re.sub(r"\s*```$", "", s)
    return s.strip()


def _as_list(v):
    if v is None:
        return []
    if isinstance(v, list):
        return v
    return [v]


def _merge_time_axes(time_axes: List[str]) -> str:
    """
    多个时间轴去重+用顿号拼接："a-b、c-d"
    """
    seen, out = set(), []
    for t in time_axes:
        t = (t or "").strip()
        if not t or t in seen:
            continue
        seen.add(t)
        out.append(t)
    return "、".join(out)


def clean_time_axis(raw: str) -> str:
    """
    清洗时间轴字符串：
    - 只保留：数字、冒号、横杠、顿号、空白
    - 去掉 '极轴' 等奇怪字符
    """
    if not raw:
        return ""
    return re.sub(r"[^\d:、\-\s]", "", str(raw))


# 匹配：……（2025-12-07 22:40:36-22:41:27） 或 ……(2025-12-07 22:40:36-22:41:27)
# 以及“2025-12-07 22:40:36-22:41:27”这种不带括号的写法
SUB_TIME_RE = re.compile(
    r"(?:[（(]\s*)?(\d{4}-\d{2}-\d{2})\s+"
    r"(\d{2}:\d{2}:\d{2}-\d{2}:\d{2}:\d{2})(?:\s*[）)])?"
)


def enrich_subclusters_with_datetime(parsed_subclusters: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    给每条子簇补字段：
    - 日期: YYYY-MM-DD
    - 时间轴: HH:MM:SS-HH:MM:SS

    抽取顺序：
    1）优先从 子簇["话题簇"] 尾部括号里的（日期+时间范围）
    2）如果没有，再尝试从 子簇["发言时间"] 里抽取（支持 "2025-12-06 22:34:24-22:38:33"）
    3）还没有，就从 _cluster_id 里取日期（时间轴留空）
    """
    out: List[Dict[str, Any]] = []

    for sc in parsed_subclusters:
        sc2 = dict(sc)

        # 1) 先看标题
        title = sc2.get("话题簇") or ""
        m = SUB_TIME_RE.search(str(title))
        if m:
            sc2["日期"] = m.group(1)
            sc2["时间轴"] = m.group(2)
        else:
            # 2) 再看发言时间
            ft = (sc2.get("发言时间") or "").strip()
            m2 = SUB_TIME_RE.search(ft)
            if m2:
                sc2["日期"] = m2.group(1)
                sc2["时间轴"] = m2.group(2)
            else:
                # 3) 兜底：日期可从 _cluster_id 里取到（时间轴取不到）
                cid = (sc2.get("_cluster_id") or "").strip()
                if cid and re.match(r"^\d{4}-\d{2}-\d{2}_", cid):
                    # 你的 _cluster_id 是 2025-12-07_B4_06，前10位就是日期
                    sc2["日期"] = cid[:10]
                sc2.setdefault("时间轴", "")

        out.append(sc2)

    return out


def normalize_model3_clusters(
    output_text: str,
    parsed_subclusters: List[Dict[str, Any]],
) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    规范化模型#3（日话题簇聚合）输出：

    输入：
      - output_text：模型#3 输出原始字符串
      - parsed_subclusters：子话题簇列表（模型#2 → aggregate 后）

    输出：
      - normalized：整理好的聚合簇列表（每个都有 话题簇 / 子话题簇列表 / 日期 / 时间轴）
      - parsed_subclusters：补完日期/时间轴后的子簇列表
    """

    # 0）先给子簇补齐 日期 + 时间轴（从标题/发言时间/_cluster_id 里扒）
    parsed_subclusters = enrich_subclusters_with_datetime(parsed_subclusters)

    # 1）建索引：_cluster_id -> 子簇（含日期/时间轴）
    sub_by_id: Dict[str, Dict[str, Any]] = {}
    for sc in parsed_subclusters:
        cid = (sc.get("_cluster_id") or "").strip()
        if cid:
            sub_by_id[cid] = sc

    # 2）“安全解析”模型#3 输出
    objs = parse_jsonl_text_safe(output_text, label="模型#3聚合输出")

    normalized: List[Dict[str, Any]] = []

    # 3）逐条处理聚合簇
    for obj in objs:
        # 3.1 统一话题簇字段
        topic = (
            obj.get("话题簇")
            or obj.get("聚合话题簇")
            or obj.get("话题簇3")
            or ""
        )
        topic = str(topic).strip()

        # 3.2 拿子话题簇列表，兼容各种 key
        sub_list = (
            obj.get("子话题簇列表")
            or obj.get("子话题簇")
            or obj.get("子话题簇id列表")
            or obj.get("子话题簇ID列表")
            or []
        )
        sub_list = [str(x).strip() for x in _as_list(sub_list) if str(x).strip()]

        # 3.3 先用聚合簇自己的 日期 / 时间轴
        date = (obj.get("日期") or "").strip()
        time_axis = (obj.get("时间轴") or "").strip()

        # 3.4 如果缺，尝试从子簇回填
        if (not date) or (not time_axis):
            sub_dates: List[str] = []
            sub_axes: List[str] = []

            for cid in sub_list:
                sc = sub_by_id.get(cid)
                if not sc:
                    continue
                d = (sc.get("日期") or "").strip()
                ta = (sc.get("时间轴") or "").strip()
                if d:
                    sub_dates.append(d)
                if ta:
                    sub_axes.append(ta)

            if not date and sub_dates:
                date = sub_dates[0]         # 多个子簇有日期，这里简单取第一个
            if not time_axis and sub_axes:
                time_axis = _merge_time_axes(sub_axes)   # 多个子簇时间轴："a-b、c-d"

        # 3.5 缺关键字段的，直接丢弃，避免后面报错
        if not topic or not sub_list or not date or not time_axis:
            continue

        # 3.6 清洗时间轴：只保留数字、冒号、横杠、顿号、空格，去掉“极轴”等脏字符
        norm_time_axis = clean_time_axis(time_axis)

        normalized.append({
            "话题簇": topic,
            "子话题簇列表": sub_list,
            "日期": date,
            "时间轴": norm_time_axis,
        })

    return normalized, parsed_subclusters


################ top5 筛选相关 #################

def parse_time_range(date_str: str, range_str: str):
    """
    从 range_str 中解析一个时间段：
    - 只认里面的 HH:MM:SS 模式
    - 少于 2 个时间点就返回 (None, None)
    """
    if not range_str:
        return None, None

    # 提取所有形如 16:10:56 的时间片段
    times = re.findall(r"\d{1,2}:\d{2}:\d{2}", str(range_str))
    if len(times) < 2:
        return None, None

    start_str, end_str = times[0], times[1]

    try:
        start_dt = datetime.strptime(f"{date_str} {start_str}", "%Y-%m-%d %H:%M:%S")
        end_dt   = datetime.strptime(f"{date_str} {end_str}", "%Y-%m-%d %H:%M:%S")
        return start_dt, end_dt
    except ValueError:
        return None, None


def match_dialogs_by_time(
    messages: List[Dict[str, Any]],
    date_str: str,
    time_axis_str: str,
) -> List[Dict[str, Any]]:
    """
    根据 日期 + 时间轴，从 messages 中筛选对应的原始发言：
    - messages 里的时间字段为：发言日期 + 发言时间
    - time_axis_str 支持多段："16:10:56-16:23:00、21:00:00-21:10:00"
    """

    if not messages or not date_str:
        return []
    if not time_axis_str or not isinstance(time_axis_str, str) or not time_axis_str.strip():
        return []

    matched: List[Dict[str, Any]] = []

    # 多个时间段用 "、" 拼接
    for part in str(time_axis_str).split("、"):
        part = part.strip()
        if not part:
            continue

        start_dt, end_dt = parse_time_range(date_str, part)
        if not start_dt or not end_dt:
            continue

        for row in messages:
            # 1) 先过滤日期不一致的
            if (row.get("发言日期") or "") != date_str:
                continue

            ts = row.get("发言时间") or ""
            try:
                msg_dt = datetime.strptime(f"{date_str} {ts}", "%Y-%m-%d %H:%M:%S")
            except Exception:
                continue

            if start_dt <= msg_dt <= end_dt:
                matched.append(row)

    return matched


def extract_cluster_stats(聚合话题簇列表: List[Dict[str, Any]], 原始发言: List[str]) -> List[Dict[str, Any]]:
    """
    统计每个聚合话题簇的发言玩家数 / 发言总数。
    """
    parsed_msgs = [json.loads(line.strip()) for line in 原始发言 if line.strip()]
    results: List[Dict[str, Any]] = []

    for cluster in 聚合话题簇列表:
        date = cluster.get("日期")
        time_axis = cluster.get("时间轴")
        if not date or not time_axis:
            print(f"⚠ 聚合话题簇缺少日期或时间轴，跳过：{cluster}")
            continue

        matched = match_dialogs_by_time(parsed_msgs, date, time_axis)
        players = {msg.get("玩家ID") for msg in matched if msg.get("玩家ID")}
        result = {
            "聚合话题簇": cluster.get("话题簇") or cluster.get("聚合话题簇"),
            "子话题簇列表": cluster.get("子话题簇列表"),
            "发言玩家总数": len(players),
            "发言总数": len(matched)
        }
        results.append(result)

    return results


def compute_heat_score(U: int, M: int) -> float:
    """
    热度公式：U * sqrt(M)
    U = 发言玩家总数；M = 发言总数
    """
    if U == 0 or M == 0:
        return 0.0
    return round(U * math.sqrt(M), 2)


def extract_top5_heat_clusters(
    聚合话题簇列表: List[Dict[str, Any]],
    原始发言: List[str],
    top_k: int = 5
) -> List[Dict[str, Any]]:
    """
    对每个聚合话题簇计算热度，返回 TopK。
    """
    parsed_msgs = [json.loads(line.strip()) for line in 原始发言 if line.strip()]
    enriched: List[Dict[str, Any]] = []

    for cluster in 聚合话题簇列表:
        date = cluster.get("日期")
        time_axis = cluster.get("时间轴")
        matched = match_dialogs_by_time(parsed_msgs, date, time_axis)

        players = {msg.get("玩家ID") for msg in matched if msg.get("玩家ID")}
        U = len(players)
        M = len(matched)
        heat = compute_heat_score(U, M)

        enriched.append({
            "聚合话题簇": cluster.get("话题簇") or cluster.get("聚合话题簇") or "未知",
            "子话题簇列表": cluster.get("子话题簇列表"),
            "日期": date,
            "时间轴": time_axis,
            "发言玩家总数": U,
            "发言总数": M,
            "热度评分": heat
        })

    # 按热度评分排序，取 TopK
    enriched.sort(key=lambda x: x["热度评分"], reverse=True)
    return enriched[:top_k]


################ 添加讨论点字段（按子话题簇聚合） ################

def attach_discussion_points(
    top_clusters: List[Dict[str, Any]],
    subclusters: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    对每个聚合簇：
    - 根据 子话题簇列表(_cluster_id) 找回子话题簇
    - 用 子话题簇["核心对象/机制"]（或["讨论点"]） 作为讨论点文本
    - 聚合成多个讨论点，每个讨论点内部挂日期时间轴列表 + 子话题簇列表
    """

    # _cluster_id -> 子话题簇 row
    sub_by_id: Dict[str, Dict[str, Any]] = {}
    for row in subclusters:
        cid = row.get("_cluster_id")
        if cid:
            sub_by_id[str(cid)] = row

    result: List[Dict[str, Any]] = []

    for cluster in top_clusters:
        ids = cluster.get("子话题簇列表", []) or []

        # 用 point_text 聚合：同一个讨论点可能对应多个子簇
        point_bucket: Dict[str, Dict[str, Any]] = {}

        for cid in ids:
            sc = sub_by_id.get(str(cid))
            if not sc:
                continue

            point_text = (sc.get("核心对象/机制") or sc.get("讨论点") or "").strip()
            if not point_text:
                continue

            d = sc.get("日期") or cluster.get("日期")
            t_axis = sc.get("时间轴") or sc.get("时间范围") or sc.get("日期时间轴") or ""

            if point_text not in point_bucket:
                point_bucket[point_text] = {
                    "讨论点": point_text,
                    "日期时间轴列表": [],
                    "子话题簇列表": [],
                }

            point_bucket[point_text]["子话题簇列表"].append(str(cid))
            point_bucket[point_text]["日期时间轴列表"].append({
                "日期": d,
                "时间轴": t_axis
            })

        # 不排序、不截断：全部返回
        points = list(point_bucket.values())

        cluster_name = cluster.get("话题簇") or cluster.get("聚合话题簇") or "未知"

        enriched_cluster = {
            "聚合话题簇": cluster_name,
            "日期": cluster.get("日期"),
            "时间轴": cluster.get("时间轴"),
            "发言玩家总数": cluster.get("发言玩家总数"),
            "发言总数": cluster.get("发言总数"),
            "热度评分": cluster.get("热度评分"),
            "讨论点": points,
        }

        result.append(enriched_cluster)

    return result


####################### 存入每日发言 top5 ########################

from typing import Union

def append_daily_top5_to_version_jsonl(
    final_result: List[Dict[str, Any]],
    version_jsonl_path: Union[str, Path],
):
    """
    将当日 Top5 追加写入【版本累计 jsonl 文件】中（例如 VERSION_TOP5_PATH）。

    同时为每条记录补充两个字段：
    1）_idx：在 version_jsonl_path 中的全局递增行号（从1开始）
    2）_daily_top_id：当天内的 Top 话题簇ID，形式：YYYY-MM-DD_TXX，例如 2025-12-02_T05

    说明：
    - 假设 final_result 内所有记录的 "日期" 相同（即同一天的 top5）
    - 如果文件中已存在同一日期的数据，会在原有基础上继续累加 _daily_top_id 的编号
    - 不会修改任何已有的 `_cluster_id` 字段
    """
    if not final_result:
        print("⚠ final_result 为空，今日无 Top5 可写入。")
        return

    date_str = (final_result[0].get("日期") or "").strip()
    if not date_str:
        raise ValueError("final_result 中缺少 '日期' 字段，无法生成 _daily_top_id。")

    path = Path(version_jsonl_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    print("✅ writing daily_top5 to:", path)
    print("✅ cwd:", Path.cwd().resolve())

    global_max_idx = 0
    existing_count_for_day = 0

    if path.exists():
        with path.open("r", encoding="utf-8") as f:
            for line_no, line in enumerate(f, start=1):
                s = line.strip()
                if not s:
                    continue
                try:
                    obj = json.loads(s)
                except json.JSONDecodeError:
                    print(f"⚠ 已有文件第{line_no}行 JSON 无法解析，已跳过。")
                    continue

                # 全局 _idx：优先用已有 _idx，若无则用行号兜底
                if isinstance(obj.get("_idx"), int):
                    global_max_idx = max(global_max_idx, obj["_idx"])
                else:
                    global_max_idx = max(global_max_idx, line_no)

                # 统计当天已有多少条（用于 _daily_top_id 编号续上去）
                if (obj.get("日期") or "").strip() == date_str:
                    existing_count_for_day += 1

    # 追加写入
    with path.open("a", encoding="utf-8") as f:
        for offset, row in enumerate(final_result, start=1):
            row = dict(row)  # 避免修改原对象引用

            # 1) 全局递增 _idx
            global_max_idx += 1
            row["_idx"] = global_max_idx

            # 2) 当天内递增 _daily_top_id
            idx_for_day = existing_count_for_day + offset
            row["_daily_top_id"] = f"{date_str}_T{idx_for_day:02d}"

            f.write(json.dumps(row, ensure_ascii=False) + "\n")

        # 可选：强制落盘
        f.flush()
        import os
        os.fsync(f.fileno())

    print(f"✅ 已将当日 Top5（含 _idx 和 _daily_top_id）追加写入: {path}")



def extract_time_axis_from_title(title: str) -> str:
    """
    从话题簇标题中提取时间轴部分，如：
    '维护补偿诉求讨论（2025-11-19 14:21:57-14:34:08）'
    或 '维护补偿诉求讨论(2025-11-19 14:21:57-14:34:08)'
    -> '14:21:57-14:34:08'
    """
    if not title:
        return ""
    title = str(title)

    # 支持全角/半角括号：（）、()
    pattern = r"[（(](\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2}-\d{2}:\d{2}:\d{2})[）)]"
    m = re.search(pattern, title)
    if m:
        return m.group(2)
    return ""


# ------- 解析模型#4的输出文本 ----------------

def parse_and_normalize_opinion_output(
    opinion_output: str,
    topic_id: str,
    discussion_point: str,
) -> Optional[Dict[str, Any]]:
    """
    解析模型#4输出，并做轻度规范化。
    前提：opinion_output 是一个单独的 JSON 对象字符串。
    """

    if not opinion_output or not opinion_output.strip():
        return None

    s = opinion_output.strip()

    # 1）最直接情况：整段就是一个 JSON 对象
    try:
        obj = json.loads(s)
    except Exception:
        # 2）兜底：如果上游哪天又在前后加了解释文字，
        #    尝试截取第一对 { ... } 中间的内容再解析一次
        start = s.find("{")
        end = s.rfind("}")
        if start == -1 or end == -1 or start >= end:
            return None
        try:
            obj = json.loads(s[start: end + 1])
        except Exception:
            return None

    # —— 用我们自己传进来的信息覆盖，确保对得上 top5_results —— 
    obj["话题簇ID"] = topic_id          # 通常是 _cluster_id，比如 "2025-12-02_B2_01"
    obj["讨论点"] = discussion_point    # 上游“核心对象/机制”字段

    # —— 统一“代表性玩家发言示例”为 list[str] —— 
    ex = obj.get("代表性玩家发言示例")
    if isinstance(ex, str):
        obj["代表性玩家发言示例"] = [ex.strip()] if ex.strip() else []
    elif isinstance(ex, list):
        obj["代表性玩家发言示例"] = [str(x).strip() for x in ex if str(x).strip()]
    else:
        obj["代表性玩家发言示例"] = []

    # —— 分歧点做个 strip，方便后面判断“无明显分歧” —— 
    diff = obj.get("玩家主要分歧点")
    if isinstance(diff, str):
        obj["玩家主要分歧点"] = diff.strip()

    return obj


# -------------- final_top5_result ---------------------

def build_daily_top5_opinion_records(
    top5_results: List[Dict[str, Any]],
    sub_opinion_map: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    把 子话题簇观点结果（sub_opinion_map） 合并回 当日top5 结果。
    """
    final_records: List[Dict[str, Any]] = []

    for cluster in top5_results:
        date = cluster.get("日期")
        cluster_name = cluster.get("聚合话题簇") or cluster.get("话题簇") or "未知"

        base = {
            "日期": date,
            "聚合话题簇": cluster_name,
            "热度评分": cluster.get("热度评分"),
            "发言玩家数": cluster.get("发言玩家总数"),
            "发言总数": cluster.get("发言总数"),
            "时间轴": cluster.get("时间轴"),
        }

        cid_list = cluster.get("子话题簇列表", []) or []

        discussion_items: List[Dict[str, Any]] = []
        all_examples: List[str] = []
        seen_examples = set()  # 防止同一句原话重复

        for idx, cid in enumerate(cid_list, start=1):
            op = sub_opinion_map.get(cid)
            if not op:
                print(f"⚠ 子话题簇 {cid} 未在 sub_opinion_map 中找到模型#4结果，跳过。")
                continue

            # 动态字段名：讨论点1 / 讨论点2 / ...
            key_discussion = f"讨论点{idx}"

            item: Dict[str, Any] = {
                key_discussion: op.get("讨论点", ""),
                "玩家共识": op.get("玩家共识", ""),
            }

            # 只在【有真实分歧】时写入“玩家主要分歧点”
            diff = op.get("玩家主要分歧点", "")
            if isinstance(diff, str):
                diff_clean = diff.strip()
                if diff_clean and diff_clean != "无明显分歧":
                    item["玩家主要分歧点"] = diff_clean

            discussion_items.append(item)

            # 收集典型发言
            examples = op.get("代表性玩家发言示例", [])
            if isinstance(examples, list):
                for ex in examples:
                    if not ex:
                        continue
                    s = str(ex)
                    if s in seen_examples:
                        continue
                    seen_examples.add(s)
                    all_examples.append(s)

        record = {
            **base,
            "讨论点列表": discussion_items,
            "代表性玩家发言示例": all_examples,
        }
        final_records.append(record)

    return final_records


# ------------------------ 版本聚合话题簇引入 -------------------------

def read_jsonl_file(path: Path) -> List[Dict[str, Any]]:
    """
    读取 jsonl 文件，返回 List[dict]。
    解析失败的行会打印告警但不会中断。
    """
    rows: List[Dict[str, Any]] = []
    if not path.exists():
        return rows

    with path.open("r", encoding="utf-8") as f:
        for i, line in enumerate(f, start=1):
            s = line.strip()
            if not s:
                continue
            try:
                obj = json.loads(s)
            except Exception as e:
                print(f"⚠ read_jsonl_file: 第{i}行 JSON 解析失败：{e} | 原文前200字符：{s[:200]}")
                continue
            rows.append(obj)
    return rows
##### 多日输出提纯，只要 id、日期、话题簇、讨论点 ################################


def _extract_points_min(points_raw: Any, max_points: int = 3) -> List[Dict[str, Any]]:
    """
    输出点结构统一为：
    {"t": "...", "tid": ["...","..."]}
    """
    if not points_raw:
        return []

    pts: List[Dict[str, Any]] = []

    if isinstance(points_raw, list):
        for p in points_raw:
            if isinstance(p, dict):
                # 兼容老字段 & 新字段
                t = (p.get("讨论点") or p.get("point") or p.get("t") or "").strip()

                tid = (
                    p.get("子话题簇列表")
                    or p.get("子话题簇id列表")
                    or p.get("subcluster_ids")
                    or p.get("tid")
                    or []
                )

                if isinstance(tid, str):
                    tid = [tid]
                if not isinstance(tid, list):
                    tid = []

                if t:
                    pts.append({"t": t, "tid": tid})

            elif isinstance(p, str):
                s = p.strip()
                if s:
                    pts.append({"t": s, "tid": []})

    elif isinstance(points_raw, str):
        s = points_raw.strip()
        if s:
            pts.append({"t": s, "tid": []})

    if max_points and len(pts) > max_points:
        pts = pts[:max_points]
    return pts



def build_version_agg_input_jsonl_text(
    daily_top5_rows: List[Dict[str, Any]],
    max_points_per_row: int = 3,
) -> str:
    """
    ✅只输出最小字段：聚合话题簇、讨论点、子话题簇列表
    每行结构：
    {"聚合话题簇":"...", "讨论点":[{"讨论点":"...","子话题簇列表":[...]}]}
    """
    out_lines: List[str] = []
    seen_ids = set()

    for r in daily_top5_rows:
        # 用于内部去重（不输出）
        _uniq = (r.get("_daily_top_id") or r.get("id") or "").strip()
        if not _uniq:
            # 兜底：用 日期+topic+_idx 做去重key（仍不输出）
            date = (r.get("日期") or "").strip()
            topic_tmp = (
                r.get("聚合话题簇")
                or r.get("话题簇")
                or r.get("话题簇3")
                or r.get("话提簇")
                or ""
            ).strip()
            idx = r.get("_idx")
            _uniq = f"{date}|{topic_tmp}|{idx}"

        if _uniq in seen_ids:
            continue
        seen_ids.add(_uniq)

        topic = (
            r.get("聚合话题簇")
            or r.get("话题簇")
            or r.get("话题簇3")
            or r.get("话提簇")
            or ""
        ).strip()

        points = _extract_points_min(r.get("讨论点"), max_points=max_points_per_row)

        # 必须有 topic & points，否则跳过
        if not topic or not points:
            continue

        obj = {
            "聚合话题簇": topic,
            "讨论点": points,
        }
        out_lines.append(json.dumps(obj, ensure_ascii=False))

    return "\n".join(out_lines)




#################### 版本热度 TopK 计算 #################

def compute_version_heat_topk(
    version_clusters: List[Dict[str, Any]],
    daily_top5_rows: List[Dict[str, Any]],
    top_k: int = 5,
) -> List[Dict[str, Any]]:
    """
    version_clusters: 模型#4 的版本聚合输出（每行有 话题簇 / 讨论点 / 日期时间轴列表）
    daily_top5_rows:  daily_top5.jsonl 解析结果（有 发言总数 / 发言玩家总数 / _daily_top_id）

    版本级发言热度公式：
      版本_发言热度 = 版本_发言玩家总数 × sqrt(版本_发言总数)
    其中：
      版本_发言总数     = 该版本级簇下所有 daily 子簇 发言总数 之和
      版本_发言玩家总数 = 该版本级簇下所有 daily 子簇 发言玩家总数 之和（近似）
    """

    # 1) 建 daily_top5 的索引：id -> {发言总数, 发言玩家总数}
    id2metrics: Dict[str, Dict[str, int]] = {}
    for r in daily_top5_rows:
        _id = (r.get("_daily_top_id") or r.get("id") or "").strip()
        if not _id:
            continue
        total_msgs = int(r.get("发言总数") or 0)
        total_players = int(r.get("发言玩家总数") or 0)
        id2metrics[_id] = {
            "发言总数": total_msgs,
            "发言玩家总数": total_players,
        }

    enriched: List[Dict[str, Any]] = []

    # 2) 对每个版本级聚合话题簇，累加旗下所有 id 的 U / M
    for vc in version_clusters:
        dt_list = vc.get("日期时间轴列表") or []
        version_total_msgs = 0
        version_total_players = 0
        used_ids = set()

        for item in dt_list:
            did = (item.get("id") or "").strip()
            if not did or did in used_ids:
                continue
            used_ids.add(did)

            metrics = id2metrics.get(did)
            if not metrics:
                # daily_top5.jsonl 里找不到这个 id，可以打印出来排查
                # print(f"[WARN] 找不到 daily 记录：{did}")
                continue

            version_total_msgs += metrics["发言总数"]
            version_total_players += metrics["发言玩家总数"]

        # 没有任何有效数据，跳过
        if version_total_msgs <= 0 or version_total_players <= 0:
            continue

        # 🔁 这里直接复用你已有的 compute_heat_score
        version_heat = compute_heat_score(version_total_players, version_total_msgs)

        vc_enriched = dict(vc)
        vc_enriched["版本_发言总数"] = version_total_msgs
        vc_enriched["版本_发言玩家总数"] = version_total_players
        vc_enriched["版本_发言热度"] = version_heat

        enriched.append(vc_enriched)

    # 3) 按版本_发言热度排序，取 TopK
    enriched.sort(key=lambda x: x.get("版本_发言热度", 0.0), reverse=True)
    return enriched[:top_k]
#####打平话提簇聚合#############

def clusters_list_to_jsonl(version_clusters: List[Dict[str, Any]]) -> str:
    """
    只把外层 list 打平成 jsonl：
    - 每行一个 {"话题簇":..., "讨论点":[...]}
    - 不拆开讨论点
    """
    lines = []
    for obj in version_clusters:
        lines.append(json.dumps(obj, ensure_ascii=False))
    return "\n".join(lines)

#########版本话提簇根据tid找到对应时间轴##########
def build_tid_time_index(daily_top5_rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    """
    从 daily_top5 记录中构建:
      tid -> [{"日期": "...", "时间轴": "..."}, ...]
    方便后续通过子话题簇 ID 回溯到具体时间段。
    """
    tid_index: Dict[str, List[Dict[str, str]]] = defaultdict(list)

    for row in daily_top5_rows:
        row_date = (row.get("日期") or "").strip()
        row_axis = (row.get("时间轴") or "").strip()
        points = row.get("讨论点") or []

        for p in points:
            tids = p.get("子话题簇列表") or p.get("tid") or []
            if isinstance(tids, str):
                tids = [tids]
            if not isinstance(tids, list):
                continue

            dt_list = p.get("日期时间轴列表") or []
            if not isinstance(dt_list, list):
                dt_list = []

            for tid in tids:
                tid = str(tid).strip()
                if not tid:
                    continue

                # 优先用 讨论点 自己的日期时间轴；没有就用 row 的日期+时间轴兜底
                if dt_list:
                    for seg in dt_list:
                        d = (seg.get("日期") or row_date).strip()
                        axis = (seg.get("时间轴") or "").strip()
                        if d and axis:
                            tid_index[tid].append({"日期": d, "时间轴": axis})
                else:
                    if row_date and row_axis:
                        tid_index[tid].append({"日期": row_date, "时间轴": row_axis})

    # 去重
    for tid, segs in tid_index.items():
        seen = set()
        uniq = []
        for seg in segs:
            key = (seg["日期"], seg["时间轴"])
            if key in seen:
                continue
            seen.add(key)
            uniq.append(seg)
        tid_index[tid] = uniq

    return tid_index
#################对单个版本话题簇：算 U/M/热度 + 生成讨论点info##############
import json
from typing import Set, List, Dict, Any

def compute_version_cluster_heat_and_points(
    cluster: Dict[str, Any],
    tid_time_index: Dict[str, List[Dict[str, str]]],
    raw_jsonl_lines: List[str],
) -> Dict[str, Any]:
    """
    针对一个“版本聚合话题簇”（模型#4 + #5 输出中的一条）：
      - 对其下每个【讨论点】单独计算热度：
          tid -> 日期+时间轴 -> 原文消息 -> U_point、M_point、heat_point
      - 话题簇整体热度 = 所有讨论点热度之和：
          发言玩家总数 = Σ U_point
          发言总数     = Σ M_point
          热度评分     = Σ heat_point
      - 输出结构按你指定：
        {
          "聚合话题簇": ...,
          "讨论点info": [{"t": "...", "日期": "...", "时间轴": "..."}, ...],
          "发言玩家总数": U_sum,
          "发言总数": M_sum,
          "热度评分": heat_sum
        }

    注意：这里不会跨讨论点去重玩家/消息：
      同一玩家在多个讨论点发言，会在不同讨论点里分别计入，然后在簇级别做“求和”。
    """
    discussion_points = cluster.get("讨论点") or []

    # 解析原始发言
    messages = [json.loads(line.strip()) for line in raw_jsonl_lines if line.strip()]

    # —— 簇级别的累积量 —— 
    total_U = 0
    total_M = 0
    total_heat = 0.0

    # 输出的讨论点info
    discussion_info: List[Dict[str, str]] = []

    for p in discussion_points:
        t_text = (p.get("t") or "").strip()
        if not t_text:
            continue

        # 该讨论点下的 tid 列表
        tid_list = p.get("tid") or []
        if isinstance(tid_list, str):
            tid_list = [tid_list]
        if not isinstance(tid_list, list):
            continue

        # 这个讨论点对应的所有 日期+时间轴 片段
        segs_for_point = []
        for tid in tid_list:
            tid = str(tid).strip()
            if not tid:
                continue
            segs_for_point.extend(tid_time_index.get(tid, []))

        # 去重这个讨论点内的 日期+时间轴 组合
        seen_seg_point = set()
        uniq_segs_point: List[Dict[str, str]] = []
        for seg in segs_for_point:
            key = (seg["日期"], seg["时间轴"])
            if key in seen_seg_point:
                continue
            seen_seg_point.add(key)
            uniq_segs_point.append(seg)

        # —— 针对这个讨论点，回溯原文，统计 U_point / M_point —— 
        matched_msgs_point: List[Dict[str, Any]] = []
        seen_msg_ids_point: Set[Any] = set()

        for seg in uniq_segs_point:
            date_str = seg["日期"]
            time_axis_str = seg["时间轴"]

            seg_msgs = match_dialogs_by_time(messages, date_str, time_axis_str)
            for msg in seg_msgs:
                # 在【讨论点内部】去重消息，避免同一消息因多段时间重复计数
                key = msg.get("_idx")
                if key is None:
                    key = (
                        msg.get("发言日期"),
                        msg.get("发言时间"),
                        msg.get("玩家ID") or msg.get("发言人ID") or msg.get("角色ID"),
                        msg.get("内容") or msg.get("文本") or msg.get("消息内容"),
                    )
                if key in seen_msg_ids_point:
                    continue
                seen_msg_ids_point.add(key)
                matched_msgs_point.append(msg)

        # 计算该讨论点的 U_point / M_point / heat_point
        players_point: Set[str] = set()
        for msg in matched_msgs_point:
            pid = msg.get("玩家ID") or msg.get("发言人ID") or msg.get("角色ID")
            if pid:
                players_point.add(str(pid))

        U_point = len(players_point)
        M_point = len(matched_msgs_point)
        heat_point = compute_heat_score(U_point, M_point)

        # 累加到簇级别
        total_U += U_point
        total_M += M_point
        total_heat += heat_point

        # 生成讨论点info（按你指定格式，只要 t + 日期 + 时间轴）
        for seg in uniq_segs_point:
            item = {
                "t": t_text,
                "日期": seg["日期"],
                "时间轴": seg["时间轴"],
            }
            discussion_info.append(item)

    # 去重讨论点info
    seen_info = set()
    uniq_info = []
    for item in discussion_info:
        key = (item["t"], item["日期"], item["时间轴"])
        if key in seen_info:
            continue
        seen_info.add(key)
        uniq_info.append(item)

    result = {
        "聚合话题簇": cluster.get("话题簇") or cluster.get("聚合话题簇") or "未知",
        "讨论点info": uniq_info,
        "发言玩家总数": total_U,
        "发言总数": total_M,
        "热度评分": round(total_heat, 2),   # 保持两位小数风格
    }

    return result
###### 对单个话题簇计算讨论点 做量化数据计算######
import json
from typing import List, Dict, Any, Set
def compute_cluster_point_metrics(
    cluster: Dict[str, Any],
    tid_time_index: Dict[str, List[Dict[str, str]]],
    raw_jsonl_lines: List[str],
) -> Dict[str, Any]:
    """
    输入：
      - cluster: 单个版本聚合话题簇（模型#4 或 #5 输出的一条）
        允许 "讨论点" 字段是：
          * list[dict]，如 [{"t": "...", "tid": [...]}, ...]
          * list[str]，如 ["讨论点1", "讨论点2"]
          * 或两者混合
    输出：
      {
        "聚合话题簇": "...",
        "讨论点info": [...],
      }
    """
    # ⭐ 统一把 cluster["讨论点"] 规范化成 [{"t": "...", "tid": [...]}, ...]
    raw_points = cluster.get("讨论点")
    # max_points=0 -> 不截断数量，保留全部讨论点
    discussion_points = _extract_points_min(raw_points, max_points=0)

    # 解析原始发言
    messages = [json.loads(line.strip()) for line in raw_jsonl_lines if line.strip()]

    discussion_info: List[Dict[str, Any]] = []

    for p in discussion_points:
        t_text = (p.get("t") or "").strip()
        if not t_text:
            continue

        # 该讨论点下的 tid 列表（允许 str / list）
        tid_list = p.get("tid") or []
        if isinstance(tid_list, str):
            tid_list = [tid_list]
        if not isinstance(tid_list, list):
            continue

        # 后面逻辑保持不变 ↓
        # 收集该讨论点对应的所有 日期+时间轴 片段
        segs_for_point: List[Dict[str, str]] = []
        for tid in tid_list:
            tid = str(tid).strip()
            if not tid:
                continue
            segs_for_point.extend(tid_time_index.get(tid, []))

        # 去重：同一讨论点内部的 日期+时间轴 组合
        seen_seg = set()
        uniq_segs: List[Dict[str, str]] = []
        for seg in segs_for_point:
            key = (seg["日期"], seg["时间轴"])
            if key in seen_seg:
                continue
            seen_seg.add(key)
            uniq_segs.append(seg)

        if not uniq_segs:
            continue

        # 回溯原文 + 统计 U/M（保持你原来的逻辑）
        matched_msgs_point: List[Dict[str, Any]] = []
        seen_msg_ids_point: Set[Any] = set()

        for seg in uniq_segs:
            date_str = seg["日期"]
            time_axis_str = seg["时间轴"]

            seg_msgs = match_dialogs_by_time(messages, date_str, time_axis_str)
            for msg in seg_msgs:
                key = msg.get("_idx")
                if key is None:
                    key = (
                        msg.get("发言日期"),
                        msg.get("发言时间"),
                        msg.get("玩家ID") or msg.get("发言人ID") or msg.get("角色ID"),
                        msg.get("内容") or msg.get("文本") or msg.get("消息内容"),
                    )
                if key in seen_msg_ids_point:
                    continue
                seen_msg_ids_point.add(key)
                matched_msgs_point.append(msg)

        players_point: Set[str] = set()
        for msg in matched_msgs_point:
            pid = msg.get("玩家ID") or msg.get("发言人ID") or msg.get("角色ID")
            if pid:
                players_point.add(str(pid))

        U_point = len(players_point)
        M_point = len(matched_msgs_point)
        heat_point = compute_heat_score(U_point, M_point)

        dates, axes = [], []
        seen_date, seen_axis = set(), set()
        for seg in uniq_segs:
            d = seg["日期"]
            a = seg["时间轴"]
            if d and d not in seen_date:
                seen_date.add(d)
                dates.append(d)
            if a and a not in seen_axis:
                seen_axis.add(a)
                axes.append(a)

        date_str_agg = "、".join(dates)
        axis_str_agg = "、".join(axes)

        info_item = {
            "t": t_text,
            "tid": tid_list,
            "日期": date_str_agg,
            "时间轴": axis_str_agg,
            "日期时间轴列表": uniq_segs,
            "发言玩家总数": U_point,
            "发言总数": M_point,
            "热度评分": heat_point,
        }
        discussion_info.append(info_item)

    return {
        "聚合话题簇": cluster.get("话题簇") or cluster.get("聚合话题簇") or "未知",
        "讨论点info": discussion_info,
    }

####2.2 对所有版本话题簇批量计算######
def compute_all_clusters_point_metrics(
    version_clusters: List[Dict[str, Any]],
    daily_top5_rows: List[Dict[str, Any]],
    raw_jsonl_lines: List[str],
) -> List[Dict[str, Any]]:
    """
    对所有版本聚合话题簇，计算其下每个讨论点的：
      - 发言玩家总数（去重）
      - 发言总数（不按内容去重）
      - 热度评分（U * sqrt(M)）

    返回：
      [
        {
          "聚合话题簇": "...",
          "讨论点info": [
            {
              "t": "...",
              "日期": "...",          # 多日期拼接
              "时间轴": "a-b、c-d",    # 多段时间拼接
              "发言玩家总数": ...,
              "发言总数": ...,
              "热度评分": ...
            },
            ...
          ]
        },
        ...
      ]
    """
    # 1) 构建 tid -> 日期时间轴 索引（来自 daily_top5）
    tid_time_index = build_tid_time_index(daily_top5_rows)

    results: List[Dict[str, Any]] = []
    for cluster in version_clusters:
        metrics = compute_cluster_point_metrics(
            cluster=cluster,
            tid_time_index=tid_time_index,
            raw_jsonl_lines=raw_jsonl_lines,
        )
        results.append(metrics)

    return results

#######3筛选出版本发言top5话提簇#######
from typing import List, Dict, Any

def extract_version_top5_clusters_from_point_metrics(
    clusters_with_points: List[Dict[str, Any]],
    top_k: int = 5,
) -> List[Dict[str, Any]]:
    """
    输入：compute_all_clusters_point_metrics 的结果：
      [
        {
          "聚合话题簇": "...",
          "讨论点info": [
            {
              "t": "...",
              "tid": [...],
              "日期": "...",
              "时间轴": "...",
              "日期时间轴列表": [...],
              "发言玩家总数": U_point,
              "发言总数": M_point,
              "热度评分": heat_point
            },
            ...
          ]
        },
        ...
      ]

    逻辑：
      1）对每个话题簇：把旗下所有讨论点的
         U / M / 热度评分 做求和，得到簇级总量
      2）按【簇级热度评分总和】排序，取 TopK 话题簇
      3）输出格式为：
        {
          "聚合话题簇": topic,
          "讨论点info": [...原样保留...],
          "发言玩家总数": U_sum,
          "发言总数": M_sum,
          "热度评分": heat_sum
        }
    """

    cluster_summaries: List[Dict[str, Any]] = []

    for cluster in clusters_with_points:
        topic = (
            cluster.get("聚合话题簇")
            or cluster.get("话题簇")
            or "未知"
        )
        points = cluster.get("讨论点info") or []

        total_U = 0
        total_M = 0
        total_heat = 0.0

        for p in points:
            try:
                u = int(p.get("发言玩家总数") or 0)
            except (TypeError, ValueError):
                u = 0
            try:
                m = int(p.get("发言总数") or 0)
            except (TypeError, ValueError):
                m = 0
            try:
                h = float(p.get("热度评分") or 0.0)
            except (TypeError, ValueError):
                h = 0.0

            total_U += u
            total_M += m
            total_heat += h

        cluster_summaries.append({
            "聚合话题簇": topic,
            "讨论点info": points,              # ⭐ 保留整个讨论点列表
            "发言玩家总数": total_U,
            "发言总数": total_M,
            "热度评分": round(total_heat, 2),   # 簇级总热度（所有讨论点热度之和）
        })

    # 按簇级热度排序，取 TopK
    cluster_summaries.sort(
        key=lambda x: x.get("热度评分", 0.0),
        reverse=True,
    )

    return cluster_summaries[:top_k]
#######版本top5的论点和时间拉平成rows#########
from typing import List, Dict, Any

def print_mech_time_from_version_top5(
    version_top5_clusters: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    输入：版本发言Top5结果（version_top5_clusters）
    输出：每个讨论点一行，只保留：
      {
        "讨论点": "...",
        "日期时间轴列表": [
          {"日期": "YYYY-MM-DD", "时间轴": "HH:MM:SS-HH:MM:SS"},
          ...
        ]
      }
    """

    rows: List[Dict[str, Any]] = []

    for cluster in version_top5_clusters:
        points = cluster.get("讨论点info") or []
        for p in points:
            t_text = (p.get("t") or "").strip()
            if not t_text:
                continue

            # 优先用已经算好的 日期时间轴列表
            dt_list = p.get("日期时间轴列表") or []

            # 如果模型输出里暂时没有这个字段，就用 日期 + 时间轴 兜底拼一条
            if not dt_list:
                date_str = (p.get("日期") or "").strip()
                axis_str = (p.get("时间轴") or "").strip()
                if date_str and axis_str:
                    dt_list = [{"日期": date_str, "时间轴": axis_str}]

            # 完全没有时间信息就跳过
            if not dt_list:
                continue

            row = {
                "讨论点": t_text,
                "日期时间轴列表": dt_list,
            }
            rows.append(row)

    return rows
######根据讨论点的时间轴提取原文#######
def get_dialogs_lines_by_dt_list_debug(
    raw_jsonl_lines: List[str],
    dt_list: List[Dict[str, str]],
    debug: bool = False,
) -> List[str]:
    """
    输入：
      - raw_jsonl_lines：原始发言 jsonl 列表（每行一个 JSON）
      - dt_list：形如 [{"日期":"2025-12-03","时间轴":"10:14:51-10:16:19"}, ...]
    输出：
      - dialogs_lines：匹配到的原文 jsonl 行列表（去重后）
    """

    # 1) 解析原始 jsonl 为 dict
    messages: List[Dict[str, Any]] = []
    for line in raw_jsonl_lines:
        s = (line or "").strip()
        if not s:
            continue
        try:
            obj = json.loads(s)
        except Exception:
            continue
        messages.append(obj)

    all_msgs: List[Dict[str, Any]] = []
    seen_msg_ids: Set[Any] = set()

    # 2) 逐个时间片调用 match_dialogs_by_time
    for seg in dt_list:
        date_str = (seg.get("日期") or "").strip()
        axis_str = (seg.get("时间轴") or "").strip()
        if not date_str or not axis_str:
            continue

        seg_msgs = match_dialogs_by_time(messages, date_str, axis_str)

        if debug:
            print(f"🔎 时间片 {date_str} {axis_str} 命中 {len(seg_msgs)} 条消息")

        for msg in seg_msgs:
            # 优先使用 _idx 去重
            key = msg.get("_idx")
            if key is None:
                key = (
                    msg.get("发言日期"),
                    msg.get("发言时间"),
                    msg.get("玩家ID") or msg.get("发言人ID") or msg.get("角色ID"),
                    msg.get("玩家消息") or msg.get("内容") or msg.get("文本") or msg.get("消息内容"),
                )

            if key in seen_msg_ids:
                continue
            seen_msg_ids.add(key)
            all_msgs.append(msg)

    # 3) 按时间排序（可选，但一般很有用）
    def _time_key(m: Dict[str, Any]):
        d = m.get("发言日期") or ""
        t = m.get("发言时间") or ""
        try:
            return datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M:%S")
        except Exception:
            return datetime.max

    all_msgs.sort(key=_time_key)

    # 4) 转回 jsonl 行
    dialogs_lines: List[str] = [json.dumps(m, ensure_ascii=False) for m in all_msgs]

    if debug:
        print(f"✅ 共合并出 {len(dialogs_lines)} 条消息（已去重）")

    return dialogs_lines
#################版本输出拆分###########
#------------------------模型4输出拆分---------------------------------

def parse_opinion_output_to_list(opinion_output: str) -> List[Dict[str, Any]]:
    """
    把模型4返回的字符串解析成 List[dict]：
    - 支持：
      * 单个 JSON 对象
      * JSON 数组
      * jsonl（多行，每行一个 JSON）
      * ```json 代码块 + 多行漂亮 JSON
    """
    if not opinion_output:
        return []

    s = opinion_output.strip()

    # 0) 去掉 ```json / ``` 外壳
    s = re.sub(r"^```[a-zA-Z0-9]*\s*", "", s)
    s = re.sub(r"\s*```$", "", s)
    s = s.strip()

    # 1) 先尝试整体解析（能吃掉单对象 / 数组）
    try:
        obj = json.loads(s)
        if isinstance(obj, dict):
            return [obj]
        if isinstance(obj, list):
            return [x for x in obj if isinstance(x, dict)]
    except json.JSONDecodeError:
        pass

    # 2) 用“括号深度”方式，从多行里拼出一个或多个 JSON 对象
    objs: List[Dict[str, Any]] = []
    buf: List[str] = []
    depth = 0

    for raw in s.splitlines():
        line = (raw or "").strip()
        if not line:
            continue

        # 去掉 markdown 列表前缀，如 "- { ... }"
        if line.startswith("- "):
            line = line[2:].lstrip()

        # 如果这一行完全不含 { 或 }，且当前深度为 0，基本是解释文字，跳过
        if depth == 0 and "{" not in line:
            continue

        # 进入/继续一个对象
        open_cnt = line.count("{")
        close_cnt = line.count("}")

        if depth == 0 and "{" in line:
            # 新对象的开始
            buf = [line]
            depth = open_cnt - close_cnt
            if depth <= 0:
                # 单行对象：{...}
                text = "\n".join(buf)
                try:
                    obj = json.loads(text)
                    if isinstance(obj, dict):
                        objs.append(obj)
                except json.JSONDecodeError:
                    pass
                buf = []
                depth = 0
        else:
            # 已经在对象内部
            buf.append(line)
            depth += open_cnt - close_cnt
            if depth <= 0:
                # 对象结束
                text = "\n".join(buf)
                try:
                    obj = json.loads(text)
                    if isinstance(obj, dict):
                        objs.append(obj)
                except json.JSONDecodeError:
                    pass
                buf = []
                depth = 0

    return objs
#############补充讨论点和发言示例至版本top5###############
from typing import List, Dict, Any

def merge_version_top5_with_opinions(
    version_top5_points: List[Dict[str, Any]],
    version_opinions: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    将版本发言Top5的量化结果（version_top5_points）
    与 模型#6的讨论点观点结果（version_opinions）合并，输出为你指定的格式：

    {
      "聚合话题簇": "...",
      "日期": "2026-01-02",
      "时间轴": "19:29:32-20:00:51",
      "讨论点列表": [
        {
          "讨论点1": "...",
          "玩家观点": [...],
          "代表性玩家发言示例": [...]
        },
        {
          "讨论点2": "...",
          ...
        }
      ]
    }
    """

    # 1) 先建一个 map: 讨论点文本 -> 模型#6的观点结果
    opinion_map: Dict[str, Dict[str, Any]] = {}
    for op in version_opinions:
        key = (op.get("讨论点") or "").strip()
        if not key:
            continue
        # 如果同名讨论点多次出现，后写入的会覆盖前一个——一般你的流程里不会重复
        opinion_map[key] = op

    merged_results: List[Dict[str, Any]] = []

    # 2) 遍历每个版本 Top5 话题簇
    for cluster in version_top5_points:
        topic = (
            cluster.get("聚合话题簇")
            or cluster.get("话题簇")
            or "未知"
        )

        points = cluster.get("讨论点info") or []

        # -------- 聚合整个簇的 日期 & 时间轴（多日、多时间段用顿号拼） --------
        date_list: List[str] = []
        time_list: List[str] = []
        seen_dates = set()
        seen_times = set()

        for p in points:
            # p["日期"] 可能是 "2025-12-02" 或 "2025-12-02、2025-12-03"
            raw_date = (p.get("日期") or "").strip()
            raw_axis = (p.get("时间轴") or "").strip()

            if raw_date:
                for d in raw_date.split("、"):
                    d = d.strip()
                    if d and d not in seen_dates:
                        seen_dates.add(d)
                        date_list.append(d)

            if raw_axis:
                for a in raw_axis.split("、"):
                    a = a.strip()
                    if a and a not in seen_times:
                        seen_times.add(a)
                        time_list.append(a)

        cluster_date = "、".join(date_list)
        cluster_axis = "、".join(time_list)

        # -------- 组装 讨论点列表：讨论点1 / 讨论点2 / ... --------
        discussion_items: List[Dict[str, Any]] = []

        for idx, p in enumerate(points, start=1):
            t_text = (p.get("t") or "").strip()
            if not t_text:
                continue

            op = opinion_map.get(t_text, {})  # 找不到就给空壳

            item = {
                f"讨论点{idx}": t_text,
                "玩家观点": op.get("玩家观点") or [],
                "代表性玩家发言示例": op.get("代表性玩家发言示例") or [],
            }
            discussion_items.append(item)

        merged_cluster = {
            "聚合话题簇": topic,
            "日期": cluster_date,
            "时间轴": cluster_axis,
            "讨论点列表": discussion_items,
        }

        merged_results.append(merged_cluster)

    return merged_results

#######提取版本top5量化数据####################
from typing import Dict, Any, List

def compute_cluster_date_coverage_raw(cluster: Dict[str, Any]) -> Dict[str, Any]:
    """
    从版本Top5的一条记录里，抽取所有讨论点的日期，计算：
      - 覆盖天数
      - 起止日期
      - 覆盖日期列表（内部使用）
    """
    dates_set = set()

    points = cluster.get("讨论点info") or []
    for p in points:
        # 优先用 日期时间轴列表
        dt_list = p.get("日期时间轴列表") or []
        if isinstance(dt_list, list) and dt_list:
            for seg in dt_list:
                d = (seg.get("日期") or "").strip()
                if d:
                    dates_set.add(d)
        else:
            # 兜底：用单独的 日期 字段
            d = (p.get("日期") or "").strip()
            if d:
                dates_set.add(d)

    if not dates_set:
        return {
            "讨论覆盖天数": 0,
            "起始日期": "",
            "结束日期": "",
            "覆盖日期列表": []
        }

    sorted_dates = sorted(dates_set)
    coverage_days = len(sorted_dates)

    return {
        "讨论覆盖天数": coverage_days,
        "起始日期": sorted_dates[0],
        "结束日期": sorted_dates[-1],
        "覆盖日期列表": sorted_dates,
    }



def build_cluster_heat_summary(
    data: Union[Dict[str, Any], List[Dict[str, Any]]]
) -> Union[Dict[str, Any], List[Dict[str, Any]]]:
    """
    通用版：
    - 如果传进来的是【单条 dict】（一条版本Top5记录），返回一条汇总 dict
    - 如果传进来的是【list[dict]】（version_top5_points 整个列表），返回 list[汇总dict]

    单条输入示例（原来那种）：
      {
        "聚合话题簇": "...",
        "讨论点info": [...],
        "发言玩家总数": 82,
        "发言总数": 586,
        "热度评分": 1178.72
      }

    单条输出示例：
      {
        "聚合话题簇": topic,
        "讨论覆盖天数": "3天（2025-12-04 ~ 2025-12-07）",
        "发言玩家总数": total_u,
        "发言总量": total_m,
        "热度评分": total_heat
      }
    """

    def _summary_one(cluster: Dict[str, Any]) -> Dict[str, Any]:
        topic = (
            cluster.get("聚合话题簇")
            or cluster.get("话题簇")
            or "未知"
        )

        # —— 计算覆盖天数 & 日期范围 —— 
        cov = compute_cluster_date_coverage_raw(cluster)
        days = cov["讨论覆盖天数"]
        start = cov["起始日期"]
        end = cov["结束日期"]

        if days > 0 and start and end:
            cover_days_str = f"{days}天（{start} ~ {end}）"
        else:
            cover_days_str = "0天"

        # —— 读取簇级别的 U/M/heat（注意转成数值）—— 
        try:
            total_u = int(cluster.get("发言玩家总数") or 0)
        except (TypeError, ValueError):
            total_u = 0

        try:
            total_m = int(cluster.get("发言总数") or 0)
        except (TypeError, ValueError):
            total_m = 0

        try:
            total_heat = float(cluster.get("热度评分") or 0.0)
        except (TypeError, ValueError):
            total_heat = 0.0

        return {
            "聚合话题簇": topic,
            "讨论覆盖天数": cover_days_str,   # "3天（2025-12-04 ~ 2025-12-07）"
            "发言玩家总数": total_u,
            "发言总量": total_m,
            "热度评分": total_heat,
        }

    # ===== 分两种情况处理 =====
    if isinstance(data, list):
        # 你现在的用法：传的是 version_top5_points（list）
        return [_summary_one(c) for c in data]
    else:
        # 兼容旧用法：传单条 dict
        return _summary_one(data)

#######抽取版本top5中话提簇和讨论点信息用于热度趋势分析##############
from typing import List, Dict, Any
import json

def extract_topic_and_points_from_version_top5(
    version_top5_points: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    输入：version_top5_points（版本发言Top5完整结构）
    输出：只保留【聚合话题簇 + 讨论点info】：
      [
        {
          "聚合话题簇": "...",
          "讨论点info": [  # 原样保留
            {
              "t": "...",
              "tid": [...],
              "日期": "...",
              "时间轴": "...",
              "日期时间轴列表": [...],
              "发言玩家总数": ...,
              "发言总数": ...,
              "热度评分": ...
            },
            ...
          ]
        },
        ...
      ]
    """
    results: List[Dict[str, Any]] = []

    for cluster in version_top5_points:
        topic = (
            cluster.get("聚合话题簇")
            or cluster.get("话题簇")
            or "未知"
        )

        points_raw = cluster.get("讨论点info") or []

        results.append({
            "聚合话题簇": topic,
            "讨论点info": points_raw,   # ✅ 原样带过去
        })

    return results


def build_heat_trend_input_jsonl(
    version_top5_points: List[Dict[str, Any]]
) -> str:
    """
    给【热度趋势智能体】喂的 JSONL：
    每行一个话题簇对象：
      {"聚合话题簇": "...", "讨论点info": [...]}
    """
    cleaned = extract_topic_and_points_from_version_top5(version_top5_points)
    lines = [json.dumps(obj, ensure_ascii=False) for obj in cleaned]
    return "\n".join(lines)
######版本输出汇总#######
def merge_version_final_summary(
    merged_opinion_version: List[Dict[str, Any]],
    heated_stats: List[Dict[str, Any]],
    heat_trend_results: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    输入：
      - merged_opinion_version：模型#6 合并后的观点结果（含 讨论点列表）
      - heated_stats：每个聚合话题簇的量化统计（讨论覆盖天数 / U / M / 热度评分）
      - heat_trend_results：模型#7 一句话热度趋势结果

    输出：
      - final_version_outputs: List[dict]，每条结构为：
        {
          "话提簇标题": "...",
          "讨论热度（量化）": {
            "讨论覆盖天数": "...",
            "发言玩家总数": X,
            "发言总量": Y,
            "热度评分": Z,
            "热度趋势": "一句话"
          },
          "讨论点列表": [...模型6里的讨论点列表...]
        }
      并按 热度评分 从大到小排序。
    """

    # 1) 建索引：topic -> 观点结果
    opinion_by_topic: Dict[str, Dict[str, Any]] = {}
    for row in merged_opinion_version:
        t = (row.get("聚合话题簇") or row.get("话题簇") or "").strip()
        if t:
            opinion_by_topic[t] = row

    # 2) 建索引：topic -> 一句话趋势
    trend_by_topic: Dict[str, str] = {}
    for row in heat_trend_results:
        t = (row.get("聚合话题簇") or row.get("话题簇") or "").strip()
        if not t:
            continue
        summary = (row.get("一句话总结") or "").strip()
        trend_by_topic[t] = summary

    # 3) 以 heated_stats 为基准顺序，拼三份信息
    final_version_outputs: List[Dict[str, Any]] = []

    for stat in heated_stats:
        topic = (stat.get("聚合话题簇") or stat.get("话题簇") or "").strip()
        if not topic:
            continue

        op = opinion_by_topic.get(topic)
        if not op:
            print(f"⚠ 未找到观点结果 merged_opinion_version：{topic}")
            continue

        trend = trend_by_topic.get(topic, "")

        # 量化信息
        cover_days_str = stat.get("讨论覆盖天数", "")
        total_u = stat.get("发言玩家总数", 0)
        total_m = stat.get("发言总量", 0)
        try:
            heat_score = float(stat.get("热度评分") or 0.0)
        except (TypeError, ValueError):
            heat_score = 0.0

        # 讨论点列表：直接用模型6合并后的结构
        discussion_list = op.get("讨论点列表") or []

        final_item = {
            "话提簇标题": topic,
            "讨论热度（量化）": {
                "讨论覆盖天数": cover_days_str,
                "发言玩家总数": total_u,
                "发言总量": total_m,
                "热度评分": heat_score,
                "热度趋势": trend,
            },
            "讨论点列表": discussion_list,
        }

        final_version_outputs.append(final_item)

    # 4) 按 热度评分 从大到小排序
    final_version_outputs.sort(
        key=lambda x: x.get("讨论热度（量化）", {}).get("热度评分", 0.0),
        reverse=True,
    )

    return final_version_outputs

