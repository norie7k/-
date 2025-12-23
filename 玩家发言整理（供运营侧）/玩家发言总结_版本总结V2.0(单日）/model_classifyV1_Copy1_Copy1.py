
from __future__ import annotations
import json, time, typing as T
import pandas as pd
import requests
import math
from pathlib import Path
from typing import List, Dict, Any,Optional
import re, json, unicodedata
from datetime import datetime
import json

# --- openpyxl 样式/工具 ---
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle

# --- docx 样式/工具 ---
from docx.oxml import OxmlElement
from docx import Document
from docx.shared import Pt
from docx.oxml.ns import qn


################模型调用，出结果###################

def load_system_prompt(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Prompt file not found: {path}")
    return path.read_text(encoding="utf-8")

def build_user_prompt_filter(json_lines: T.List[str]) -> str:
    # 模型#1：筛掉非游戏相关，只输出相关 JSON 行（原样）
    return (
        "以下是若干玩家/客服/研发的发言记录，请根据系统提示中规则，"
        "判断哪些是【与游戏内容相关】的发言，保留这些 JSON 行，不相关的忽略。"
        "请仅输出【相关发言的原始 JSON 行】，严格保持格式不变。\n\n"
        "【输入】：\n" + "\n".join(json_lines)
    )

def build_user_prompt_clsuter(jsonl_block: str) -> str:
    return (
        "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请先完整阅读全部输入，然后按系统提示中的话题簇规则进行划分。\n"
        "【输出要求】只输出若干 JSON 对象，每个话题簇一个 JSON；"
        "禁止使用 ```json 或 ``` 等 Markdown 代码块，禁止输出解释文字。\n\n"
        "【输入】：\n" + jsonl_block
    )

def build_user_prompt_cluster_agg(jsonl_block: str) -> str:
    return (
        "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请先完整阅读全部输入，然后按系统提示中的话题簇规则进行划分。\n"
        "【输出要求】只输出若干 JSON 对象，每个话题簇一个 JSON；"
        "禁止使用 ```json 或 ``` 等 Markdown 代码块，禁止输出解释文字。\n\n"
        "【输入】：\n" + jsonl_block
    )
from typing import List, Dict, Any
import json

def build_user_prompt_subcluster_opinion(
    topic_id: str,
    discussion_point: str,
    dialogs: List[Dict[str, Any]],
) -> str:
    lines = []

    meta = {
        "话题簇ID": topic_id,
        "讨论点": discussion_point,
    }
    lines.append(json.dumps(meta, ensure_ascii=False))

    for row in dialogs:
        lines.append(json.dumps(row, ensure_ascii=False))

    jsonl_block = "\n".join(lines)

    return (
        "禁止使用 ```json 或 ``` 等 Markdown 代码块，禁止输出解释文字。\n\n"
        "【输入】本次所有数据（JSONL，第一行为话题簇信息，其余为发言）：\n"
        + jsonl_block +
        "\n\n仅根据以上内容，并遵循系统提示词中的规则和输出格式，"
        "直接输出 1 个 JSON 对象作为最终结果。"
    )

def call_ark_chat_completions(
    api_url: str,
    api_key: str,
    model: str,
    system_prompt: str,
    user_prompt: str,
    temperature: float = 0.3,
    max_tokens: int = 32700,
    timeout: int = 600,
    retries: int = 2,
) -> str:
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    last_err = None
    for attempt in range(retries + 1):
        try:
            resp = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text}")
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            last_err = e
            time.sleep(1.2 * (attempt + 1))
    raise RuntimeError(f"Ark API 调用失败: {last_err}")

def extract_valid_json_lines(text: str) -> T.List[str]:
    """
    把模型输出里的纯 JSON 行提取出来（鲁棒处理）：
    - 逐行判断：以 { 开头 且 以 } 结尾，则认为是一个 JSON 对象行
    - 也能容忍前后多余空行或解释文字（会被忽略）
    """
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("{") and s.endswith("}"):
            lines.append(s)
    return lines

def add_index_to_jsonl_lines(jsonl_lines):
    """
    给原始 jsonl 每条发言加上一个唯一行号字段 `_idx`，
    返回新的 List[str]，每个元素仍然是一行 JSON 字符串。
    """
    new_lines = []
    for idx, line in enumerate(jsonl_lines, start=1):
        line = line.strip()
        if not line:
            continue
        obj = json.loads(line)
        obj["_idx"] = idx  # 新增行号
        new_lines.append(json.dumps(obj, ensure_ascii=False))
    return new_lines


def count_output_filter_stats(output_filter: str):
    """
    统计模型#1 输出中的：
    - 总行数 total_lines（玩家+客服+研发）
    - 玩家发言行数 player_lines（简单用 发言人ID 里的关键词区分）
    """
    total_lines = 0
    player_lines = 0

    for line in output_filter.splitlines():
        line = line.strip()
        if not line:
            continue
        total_lines += 1

        try:
            obj = json.loads(line)
        except Exception:
            # 如果这一行不是合法 JSON，就只能算进 total_lines，无法细分
            continue

        speaker = (
            obj.get("玩家ID")
            or obj.get("发言人ID")
            or obj.get("角色ID")
            or ""
        )

        # 简单排除客服/GM/官方/运营/研发，其余视为玩家
        if any(key in str(speaker) for key in ["客服", "GM", "官方", "运营", "研发"]):
            continue

        player_lines += 1

    return total_lines, player_lines


def get_covered_indices_from_cluster_output(output_cluster: str):
    """
    从模型#2 的自然语言输出中，解析所有 “发言行号列表：[...数字...]” 里的数字，
    收集成一个 set 返回。

    适配类似格式（你现在的输出就是这样）：
        发言行号列表：[1, 2, 3]
        发言行号列表：[88, 90]
        发言行号列表：[124,125,126]

    支持：
    - 中文/英文冒号（: / ：）
    - 中文/英文中括号（[] / 【】）
    - 逗号可为 , 或 ，
    """
    covered = set()

    # 正则匹配 “发言行号列表：[1, 2, 3]”
    pattern = r"发言行号列表[:：]\s*[\[\【]([0-9,\s，]+)[\]\】]"

    for m in re.finditer(pattern, output_cluster):
        nums_str = m.group(1)  # 里面是类似 "1, 2, 3" 或 "124,125,126"
        # 替换中文逗号，按逗号切分
        nums_str = nums_str.replace("，", ",")
        for part in nums_str.split(","):
            p = part.strip()
            if not p:
                continue
            try:
                covered.add(int(p))
            except ValueError:
                continue

    return covered


#################################话提簇唯一id#################################

###1. clusterid的日期扒取##
def infer_date_for_batch(
    cluster_json_list: List[Dict[str, Any]],
    batch_lines: List[str]
) -> str:
    """
    优先从话题簇标题中抽取日期（形如：xxxx（2025-12-02 16:30:01-17:42:41））
    若失败，则回退到原始发言中的 `发言日期` / `日期` 字段。
    """
    # 1）先从话题簇标题里找 YYYY-MM-DD
    for obj in cluster_json_list:
        title = str(
            obj.get("话题簇")
            or obj.get("聚合话题簇")
            or ""
        )
        m = re.search(r"(\d{4}-\d{2}-\d{2})", title)
        if m:
            return m.group(1)

    # 2）如果标题里没有，就从原始发言里拿
    for line in batch_lines:
        line = line.strip()
        if not line:
            continue
        try:
            msg = json.loads(line)
        except Exception:
            continue

        date_str = msg.get("发言日期") or msg.get("日期")
        if date_str:
            return date_str

    # 3）再不行就报错，让你意识到数据格式有问题
    raise ValueError("无法从话题簇标题或原始发言中推断出日期，请检查数据格式。")

##2.生成话提簇clusterid##
def assign_global_cluster_ids(cluster_list, date_str, batch_id):
    """
    为每个话题簇生成全局唯一ID字段 `_cluster_id`
    格式：YYYY-MM-DD_BX_XX，如 2025-11-20_B2_03
    """
    for idx, cluster in enumerate(cluster_list, start=1):
        cluster["_cluster_id"] = f"{date_str}_{batch_id}_{idx:02d}"
    return cluster_list

#################################聚合每天的话提簇分批输出#################################
def aggregate_cluster_outputs(batch_outputs: List[str]) -> str:
    all_lines = []

    for batch_id, text in enumerate(batch_outputs, start=1):
        if not text:
            continue

        for line in text.strip().splitlines():
            line = line.strip()
            if not line:
                continue

            try:
                obj = json.loads(line)
            except Exception:
                # 如果有一行不是合法 JSON，就跳过（也可以改成 raise，看你需求）
                continue

            clean_line = json.dumps(obj, ensure_ascii=False)
            all_lines.append(clean_line)

    # 聚合为一个大的 JSONL 字符串
    return "\n".join(all_lines)


################top5筛选#################################


# -------------------------------
# 🔹 1. 匹配原始发言
# -------------------------------
from datetime import datetime
import json
from typing import List, Dict

def parse_time_range(date_str, range_str):
    # 如果只有一个时间点而不是范围，跳过
    if "-" not in range_str:
        return None, None
    start_str, end_str = range_str.split("-")
    start_dt = datetime.strptime(f"{date_str} {start_str.strip()}", "%Y-%m-%d %H:%M:%S")
    end_dt = datetime.strptime(f"{date_str} {end_str.strip()}", "%Y-%m-%d %H:%M:%S")
    return start_dt, end_dt


def match_dialogs_by_time(messages, date_str, time_axis_str):
    time_ranges = time_axis_str.split("、")
    matched = []
    for tr in time_ranges:
        start, end = parse_time_range(date_str, tr)
        if not start or not end:
            continue  # 跳过无法解析的时间段
        for row in messages:
            ts = datetime.strptime(f"{row['发言日期']} {row['发言时间']}", "%Y-%m-%d %H:%M:%S")
            if start <= ts <= end:
                matched.append(row)
    return matched

def get_time_axis(cluster: dict) -> str | None:
    # 优先标准字段
    for key in ("时间轴", "time轴", "时间轴 ", "时间段", "时间区间"):
        val = cluster.get(key)
        if isinstance(val, str) and val.strip():
            return val.strip()
    return None

def extract_cluster_stats(聚合话题簇列表: List[Dict], 原始发言: List[str]) -> List[Dict]:
    parsed_msgs = [json.loads(line.strip()) for line in 原始发言 if line.strip()]
    results = []

    for cluster in 聚合话题簇列表:
        date = cluster.get("日期")
        time_axis = get_time_axis(cluster)
        if not date or not time_axis:
            print(f"⚠ 聚合话题簇缺少日期或时间轴，跳过：{cluster}")
            continue
        matched = match_dialogs_by_time(parsed_msgs, date, time_axis)
        players = {msg.get("玩家ID") for msg in matched if msg.get("玩家ID")}
        result = {
            "聚合话题簇": cluster.get("话题簇") or cluster.get("聚合话题簇"),
            "子话题簇列表": cluster.get("子话题簇列表"),
            "发言玩家总数": len(players),
            "发言总数": len(matched)
        }
        results.append(result)

    return results

# -------------------------------
# 🔹 2. 热度识别主函数
# -------------------------------
def compute_heat_score(U: int, M: int) -> float:
    if U == 0 or M == 0:
        return 0.0
    return round(U * math.sqrt(M), 2)

def extract_top5_heat_clusters(聚合话题簇列表: List[Dict], 原始发言: List[str], top_k=5) -> List[Dict]:
    parsed_msgs = [json.loads(line.strip()) for line in 原始发言 if line.strip()]
    enriched = []

    for cluster in 聚合话题簇列表:
        date = cluster.get("日期")
        time_axis = cluster.get("时间轴")
        matched = match_dialogs_by_time(parsed_msgs, date, time_axis)

        players = {msg.get("玩家ID") for msg in matched if msg.get("玩家ID")}
        U = len(players)
        M = len(matched)
        heat = compute_heat_score(U, M)

        enriched.append({
        "聚合话题簇": cluster.get("话题簇") or cluster.get("聚合话题簇") or "未知",
        "子话题簇列表": cluster.get("子话题簇列表"),
        "日期": date,
        "时间轴": time_axis,
        "发言玩家总数": U,
        "发言总数": M,
        "热度评分": heat
         })


    # 按热度评分排序，取 TopK
    enriched.sort(key=lambda x: x["热度评分"], reverse=True)
    return enriched[:top_k]

# -------------------------------
# 🔹 3. 添加讨论点字段
# -------------------------------
def attach_discussion_points(top_clusters: List[Dict], subclusters: List[Dict]) -> List[Dict]:
    """
    将 top_clusters 中每个聚合簇的子话题簇列表，与 subclusters 中的 _cluster_id 匹配，
    拼出核心机制描述，并格式化为列点 + 空行。移除子话题簇列表字段。
    """
    # 构建 _cluster_id → 核心机制 映射
    cluster_mechanism_map = {
        row["_cluster_id"]: row["核心对象/机制"]
        for row in subclusters
        if "_cluster_id" in row and "核心对象/机制" in row
    }

    result = []
    for cluster in top_clusters:
        ids = cluster.get("子话题簇列表", [])
        mechanisms = [cluster_mechanism_map.get(cid) for cid in ids if cid in cluster_mechanism_map]

        # 格式化为列点 + 空行
        discussion_point = [m for m in mechanisms if m]

        # 修复聚合话题簇名称字段为空的问题
        cluster_name = cluster.get("话题簇") or cluster.get("聚合话题簇") or "未知"

        enriched_cluster = {
            "聚合话题簇": cluster_name,
            "日期": cluster.get("日期"),
            "时间轴": cluster.get("时间轴"),
            "发言玩家总数": cluster.get("发言玩家总数"),
            "发言总数": cluster.get("发言总数"),
            "热度评分": cluster.get("热度评分"),
            "讨论点": discussion_point
        }

        result.append(enriched_cluster)

    return result



####################### 存入每日发言 top5 ########################
from pathlib import Path
from typing import List, Dict
import json

####################### 存入每日发言 top5 ########################
from pathlib import Path
from typing import List, Dict
import json

def append_daily_top5_to_version_jsonl(
    final_result: List[Dict],
    version_jsonl_path: str = "version_daily_top5_with_opinion.jsonl",
):
    """
    将当日 Top5 追加写入某个【版本累计 jsonl 文件】中。
    
    同时为每条记录补充两个字段：
    1）_idx：在 version_jsonl_path 中的全局递增行号（从1开始）
    2）_daily_top_id：当天内的 Top 话题簇ID，形式：YYYY-MM-DD_TXX，例如 2025-12-03_T01
    
    说明：
    - 假设 final_result 内所有记录的 "日期" 相同（即同一天的 top5）
    - 如果文件中已存在同一日期的数据，会在原有基础上继续累加 _daily_top_id 的编号
    - 不会修改任何已有的 `_cluster_id` 字段（子话题簇仍然用它）
    """
    if not final_result:
        print("⚠ final_result 为空，今日无 Top5 可写入。")
        return

    # 取当天日期（假设 final_result 同一批都是同一天）
    date_str = final_result[0].get("日期")
    if not date_str:
        raise ValueError("final_result 中缺少 '日期' 字段，无法生成 _daily_top_id。")

    path = Path(version_jsonl_path)

    # ---------- 1. 计算当前文件中的最大 _idx（全局行号） ----------
    global_max_idx = 0
    # 同时统计“今天”已经有多少条，用于 _daily_top_id 续编号
    existing_count_for_day = 0

    if path.exists():
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError:
                    continue

                # 统计全局 _idx
                if isinstance(obj.get("_idx"), int):
                    global_max_idx = max(global_max_idx, obj["_idx"])
                else:
                    # 旧数据没有 _idx，就简单当作一行，新行号往后推
                    global_max_idx += 1

                # 统计当天的条数（用于 _daily_top_id）
                if obj.get("日期") == date_str:
                    existing_count_for_day += 1

    # ---------- 2. 追加写入，并为新记录生成 _idx + _daily_top_id ----------
    mode = "a" if path.exists() else "w"
    with path.open(mode, encoding="utf-8") as f:
        for offset, row in enumerate(final_result, start=1):
            row = dict(row)  # 避免修改原对象引用

            # 2.1 全局递增 _idx（版本所有天共用一个序号）
            global_max_idx += 1
            row["_idx"] = global_max_idx

            # 2.2 生成当天的 top id（不碰 `_cluster_id`）
            # 当天已有 existing_count_for_day 条，这批从 +1 开始接着排
            idx_for_day = existing_count_for_day + offset
            row["_daily_top_id"] = f"{date_str}_T{idx_for_day:02d}"

            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    print(f"✅ 已将当日 Top5（含 _idx 和 _daily_top_id）追加写入: {path}")


def extract_time_axis_from_title(title: str) -> str:
    """
    从话题簇标题中提取时间轴部分，如：
    '维护补偿诉求讨论（2025-11-19 14:21:57-14:34:08）'
    或 '维护补偿诉求讨论(2025-11-19 14:21:57-14:34:08)'
    -> '14:21:57-14:34:08'
    """
    if not title:
        return ""
    title = str(title)

    # 支持全角/半角括号：（）、()
    pattern = r"[（(](\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2}-\d{2}:\d{2}:\d{2})[）)]"
    m = re.search(pattern, title)
    if m:
        # 你的 match_dialogs_by_time 是日期+时间段分开传的，所以只要 HH:MM:SS-HH:MM:SS 即可
        return m.group(2)
    return ""
#-------解析模型#4的输出文本----------------
def parse_and_normalize_opinion_output(
    opinion_output: str,
    topic_id: str,
    discussion_point: str,
) -> Optional[Dict[str, Any]]:
    """
    解析模型#4输出，并做轻度规范化。
    前提：opinion_output 是一个单独的 JSON 对象字符串。

    返回：规范化后的 dict；
         如果解析失败则返回 None。
    """
    if not opinion_output or not opinion_output.strip():
        return None

    s = opinion_output.strip()

    # 1）最直接情况：整段就是一个 JSON 对象
    try:
        obj = json.loads(s)
    except Exception:
        # 2）兜底：如果上游哪天又在前后加了解释文字，
        #    尝试截取第一对 { ... } 中间的内容再解析一次
        start = s.find("{")
        end = s.rfind("}")
        if start == -1 or end == -1 or start >= end:
            return None
        try:
            obj = json.loads(s[start : end + 1])
        except Exception:
            return None

    # —— 用我们自己传进来的信息覆盖，确保对得上 top5_results —— 
    obj["话题簇ID"] = topic_id          # 通常是 _cluster_id，比如 "2025-12-02_B2_01"
    obj["讨论点"] = discussion_point    # 上游“核心对象/机制”字段

    # —— 统一“代表性玩家发言示例”为 list[str] —— 
    ex = obj.get("代表性玩家发言示例")
    if isinstance(ex, str):
        obj["代表性玩家发言示例"] = [ex.strip()] if ex.strip() else []
    elif isinstance(ex, list):
        obj["代表性玩家发言示例"] = [str(x).strip() for x in ex if str(x).strip()]
    else:
        obj["代表性玩家发言示例"] = []

    # —— 分歧点做个 strip，方便后面判断“无明显分歧” —— 
    diff = obj.get("玩家主要分歧点")
    if isinstance(diff, str):
        obj["玩家主要分歧点"] = diff.strip()

    return obj

#--------------final_top5_result---------------------

def build_daily_top5_opinion_records(
    top5_results: List[Dict[str, Any]],
    sub_opinion_map: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    final_records: List[Dict[str, Any]] = []

    for cluster in top5_results:
        date = cluster.get("日期")
        cluster_name = cluster.get("聚合话题簇") or cluster.get("话题簇") or "未知"

        base = {
            "日期": date,
            "聚合话题簇": cluster_name,
            "热度评分": cluster.get("热度评分"),
            "发言玩家数": cluster.get("发言玩家总数"),
            "发言总数": cluster.get("发言总数"),
            "时间轴": cluster.get("时间轴"),
        }

        cid_list = cluster.get("子话题簇列表", []) or []

        discussion_items: List[Dict[str, Any]] = []
        all_examples: List[str] = []
        seen_examples = set()  # 防止同一句原话重复

        for idx, cid in enumerate(cid_list, start=1):
            op = sub_opinion_map.get(cid)
            if not op:
                print(f"⚠ 子话题簇 {cid} 未在 sub_opinion_map 中找到模型#4结果，跳过。")
                continue

            # 动态字段名：讨论点1 / 讨论点2 / ...
            key_discussion = f"讨论点{idx}"

            item: Dict[str, Any] = {
                key_discussion: op.get("讨论点", ""),
                "玩家共识": op.get("玩家共识", ""),
            }

            # 只在【有真实分歧】时写入“玩家主要分歧点”
            diff = op.get("玩家主要分歧点", "")
            if isinstance(diff, str):
                diff_clean = diff.strip()
                if diff_clean and diff_clean != "无明显分歧":
                    item["玩家主要分歧点"] = diff_clean

            discussion_items.append(item)

            # 收集典型发言
            examples = op.get("代表性玩家发言示例", [])
            if isinstance(examples, list):
                for ex in examples:
                    if not ex:
                        continue
                    s = str(ex)
                    if s in seen_examples:
                        continue
                    seen_examples.add(s)
                    all_examples.append(s)

        record = {
            **base,
            "讨论点列表": discussion_items,
            "代表性玩家发言示例": all_examples,
        }
        final_records.append(record)

    return final_records
def normalize_agg_cluster_key_name(raw_key: str) -> str:
    """
    将模型#3输出中的各种乱七八糟key，统一映射成规范字段名。

    目标规范字段：
    - 话题簇
    - 子话题簇列表
    - 日期
    - 时间轴

    其他乱造字段保留原名（你之后如果想继续扩展，可以在这里加规则）
    """
    if not isinstance(raw_key, str):
        raw_key = str(raw_key)
    k = raw_key.strip()
    lower = k.lower()

    # 1) 话题簇本体（聚合后的名字）
    if k in ("话题簇", "聚合话题簇", "聚合簇", "聚合话提簇"):
        return "话题簇"
    # 误拼：话提簇 / 话提族
    if ("话题" in k or "话提" in k) and ("簇" in k or "族" in k) and "子" not in k:
        return "话题簇"

    # 2) 子话题簇列表
    if "子" in k and ("话题" in k or "话提" in k) and ("簇" in k or "族" in k):
        return "子话题簇列表"
    if "子簇" in k or "子话题簇id" in lower:
        return "子话题簇列表"

    # 3) 日期
    if "日期" in k or lower == "date" or "发言日期" in k:
        return "日期"

    # 4) 时间轴（各种奇怪写法：time轴 / 时间极 / 极轴 / 时间轴 带空格等）
    if "时间轴" == k:
        return "时间轴"
    if "时间" in k or "时" in k or "轴" in k or "极轴" in k or "time" in lower:
        # 这里直接收敛成 时间轴，你的数据结构里本来也只应该有一个“时间段字段”
        return "时间轴"

    # 5) 其他字段（可能是你之后算的 发言玩家总数 / 发言总数 / 热度评分 等）原样保留
    return k
def normalize_agg_cluster_object(obj: dict) -> dict:
    """
    对【一条】聚合话题簇记录做 key 归一化：
    - 把各种错别字、变体key 收敛成标准key
    - 如果多个 key 映射到同一个新key，按简单策略合并：
        - 如果都是 list，就直接拼接
        - 否则后出现的覆盖前面的（你也可以换成保留第一个）
    """
    fixed = {}
    for raw_k, v in obj.items():
        new_k = normalize_agg_cluster_key_name(raw_k)

        # 合并策略
        if new_k in fixed:
            old_v = fixed[new_k]
            if isinstance(old_v, list) and isinstance(v, list):
                fixed[new_k] = old_v + v
            else:
                # 默认用后者覆盖前者；如果你更想保留第一次结果，可以改成 `continue`
                fixed[new_k] = v
        else:
            fixed[new_k] = v

    return fixed
import json

def fix_output_cluster_agg_keys(output_cluster_agg: str) -> str:
    """
    读取模型#3返回的文本（output_cluster_agg），
    逐行解析 JSON，对每条记录的 key 做规范化处理，
    再重新 dump 回 jsonl 字符串。

    - 自动跳过非 JSON 行（比如模型在前后乱加的说明）
    - 自动跳过无法解析的脏行，并打印告警
    """
    fixed_lines = []

    for line in output_cluster_agg.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        if not line.startswith("{"):
            # 非 JSON 行统一跳过，避免 json.loads 报错
            print("⚠ 模型#3 输出中发现非 JSON 行，被跳过：", line[:80])
            continue

        try:
            obj = json.loads(line)
        except json.JSONDecodeError as e:
            print("⚠ 聚合簇 JSON 解析失败，已跳过：", e, "原始行:", line[:120])
            continue

        fixed_obj = normalize_agg_cluster_object(obj)
        fixed_lines.append(json.dumps(fixed_obj, ensure_ascii=False))

    return "\n".join(fixed_lines)
