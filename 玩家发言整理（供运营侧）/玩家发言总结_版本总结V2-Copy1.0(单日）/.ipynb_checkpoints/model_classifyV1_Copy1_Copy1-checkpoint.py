
from __future__ import annotations
import json, time, typing as T
import pandas as pd
import requests
import math
from pathlib import Path
from typing import List, Dict, Any,Optional,Union,Tuple
import re, json, unicodedata
from datetime import datetime
import json
from collections import defaultdict
from json import JSONDecodeError
# --- openpyxl 样式/工具 ---
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle

# --- docx 样式/工具 ---
from docx.oxml import OxmlElement
from docx import Document
from docx.shared import Pt
from docx.oxml.ns import qn


################模型调用，出结果###################

def load_system_prompt(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Prompt file not found: {path}")
    return path.read_text(encoding="utf-8")

def build_user_prompt_filter(json_lines: T.List[str]) -> str:
    # 模型#1：筛掉非游戏相关，只输出相关 JSON 行（原样）
    return (
        "以下是若干玩家/客服/研发的发言记录，请根据系统提示中规则，"
        "判断哪些是【与游戏内容相关】的发言，保留这些 JSON 行，不相关的忽略。"
        "请仅输出【相关发言的原始 JSON 行】，严格保持格式不变。\n\n"
        "【输入】：\n" + "\n".join(json_lines)
    )

def build_user_prompt_clsuter(jsonl_block: str) -> str:
    return (
        "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请先完整阅读全部输入，然后按系统提示中的话题簇规则进行划分。\n"
        "【输出要求】只输出若干 JSON 对象，每个话题簇一个 JSON；"
        "禁止使用 ```json 或 ``` 等 Markdown 代码块，禁止输出解释文字。\n\n"
        "【输入】：\n" + jsonl_block
    )

def build_user_prompt_cluster_agg(jsonl_block: str) -> str:
    return (
        "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请先完整阅读全部输入，然后按系统提示中的话题簇规则进行划分。\n"
        "【输出要求】只输出若干 JSON 对象，每个话题簇一个 JSON；"
        "禁止使用 ```json 或 ``` 等 Markdown 代码块，禁止输出解释文字。\n\n"
        "【输入】：\n" + jsonl_block
        
    )



from typing import List, Any

def build_user_prompt_subcluster_opinion(
    discussion_point: str,
    json_lines: List[Any],   # 实际传的是 list[str]
) -> str:
    dp = (discussion_point or "").strip()
    jsonl_block = "\n".join(
        (line or "").strip() for line in json_lines if line
    )
    return (
        "你将收到一段玩家聊天原文（JSONL，每行一条）。本次只分析指定【讨论点】，其他内容忽略。\n\n"
        "【输出要求】只输出1个JSON对象，禁止解释文字，禁止Markdown代码块。\n\n"
        "【输入话题点】：\n" + dp + "\n\n"
        "【输入原文JSONL】\n" + jsonl_block
    )

def call_ark_chat_completions(
    api_url: str,
    api_key: str,
    model: str,
    system_prompt: str,
    user_prompt: str,
    temperature: float = 0.3,
    max_tokens: int = 32700,
    timeout: int = 600,
    retries: int = 2,
) -> str:
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    last_err = None
    for attempt in range(retries + 1):
        try:
            resp = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text}")
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            last_err = e
            time.sleep(1.2 * (attempt + 1))
    raise RuntimeError(f"Ark API 调用失败: {last_err}")

def extract_valid_json_lines(text: str) -> T.List[str]:
    """
    把模型输出里的纯 JSON 行提取出来（鲁棒处理）：
    - 逐行判断：以 { 开头 且 以 } 结尾，则认为是一个 JSON 对象行
    - 也能容忍前后多余空行或解释文字（会被忽略）
    """
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("{") and s.endswith("}"):
            lines.append(s)
    return lines

def add_index_to_jsonl_lines(jsonl_lines):
    """
    给原始 jsonl 每条发言加上一个唯一行号字段 `_idx`，
    返回新的 List[str]，每个元素仍然是一行 JSON 字符串。
    """
    new_lines = []
    for idx, line in enumerate(jsonl_lines, start=1):
        line = line.strip()
        if not line:
            continue
        obj = json.loads(line)
        obj["_idx"] = idx  # 新增行号
        new_lines.append(json.dumps(obj, ensure_ascii=False))
    return new_lines


def count_output_filter_stats(output_filter: str):
    """
    统计模型#1 输出中的：
    - 总行数 total_lines（玩家+客服+研发）
    - 玩家发言行数 player_lines（简单用 发言人ID 里的关键词区分）
    """
    total_lines = 0
    player_lines = 0

    for line in output_filter.splitlines():
        line = line.strip()
        if not line:
            continue
        total_lines += 1

        try:
            obj = json.loads(line)
        except Exception:
            # 如果这一行不是合法 JSON，就只能算进 total_lines，无法细分
            continue

        speaker = (
            obj.get("玩家ID")
            or obj.get("发言人ID")
            or obj.get("角色ID")
            or ""
        )

        # 简单排除客服/GM/官方/运营/研发，其余视为玩家
        if any(key in str(speaker) for key in ["客服", "GM", "官方", "运营", "研发"]):
            continue

        player_lines += 1

    return total_lines, player_lines


def get_covered_indices_from_cluster_output(output_cluster: str):
    """
    从模型#2 的自然语言输出中，解析所有 “发言行号列表：[...数字...]” 里的数字，
    收集成一个 set 返回。

    适配类似格式（你现在的输出就是这样）：
        发言行号列表：[1, 2, 3]
        发言行号列表：[88, 90]
        发言行号列表：[124,125,126]

    支持：
    - 中文/英文冒号（: / ：）
    - 中文/英文中括号（[] / 【】）
    - 逗号可为 , 或 ，
    """
    covered = set()

    # 正则匹配 “发言行号列表：[1, 2, 3]”
    pattern = r"发言行号列表[:：]\s*[\[\【]([0-9,\s，]+)[\]\】]"

    for m in re.finditer(pattern, output_cluster):
        nums_str = m.group(1)  # 里面是类似 "1, 2, 3" 或 "124,125,126"
        # 替换中文逗号，按逗号切分
        nums_str = nums_str.replace("，", ",")
        for part in nums_str.split(","):
            p = part.strip()
            if not p:
                continue
            try:
                covered.add(int(p))
            except ValueError:
                continue

    return covered



################################修补bug话提簇划分解析#########################

def _normalize_json_text(text: str) -> str:
    """
    对模型输出做一些小修复，以便 json.loads 正常解析：
    1）去掉整段末尾多余的逗号：{"a":1},
    2）去掉 属性/元素 后紧跟 } 或 ] 的非法逗号：
        例如：
            "核心对象/机制": "xxx",
        }
        →  "核心对象/机制": "xxx"
           }
    """
    text = text.rstrip()

    # 1) 整个对象末尾多打一颗逗号：{"a":1},
    if text.endswith(","):
        text = text[:-1].rstrip()

    # 2) 属性/元素后直接跟 } 或 ] 的逗号：
    #    把 ",\n  }" 或 ",\n]" 变成 "\n  }" 或 "\n]"
    text = re.sub(r",\s*(\n\s*[}\]])", r"\1", text)

    return text

def parse_model2_output_to_json_list(output_cluster: str, batch_idx: int = 0) -> List[Dict[str, Any]]:
    """
    把模型#2的原始输出解析成 list[dict]：
    - 支持一行一个 JSON：
        {"话题簇1": "...", "核心对象/机制": "..."}
    - 也支持多行漂亮 JSON：
        {
          "话题簇1": "...",
          "核心对象/机制": "..."
        }
    - 自动忽略解释性文字、markdown 列表 "- {...}"
    - 调用 _normalize_json_text 修复尾逗号等格式问题
    - 解析失败时打印完整对象原文，但不中断整批
    """
    objs: List[Dict[str, Any]] = []
    cur_lines: List[str] = []
    depth = 0  # 当前大括号嵌套层数

    lines = output_cluster.strip().splitlines()

    for idx, raw in enumerate(lines, start=1):
        line = (raw or "").rstrip()
        if not line.strip():
            continue

        # 去掉 markdown 列表前缀（常见格式："- {...}"）
        if line.lstrip().startswith("- "):
            line = line.lstrip()[2:].lstrip()

        # 当前还没进入任何 JSON 对象，并且这一行也看不到 "{"
        if depth == 0 and "{" not in line:
            # 大概率是解释性文字，比如 "以下是话题簇..."
            # 如果想看，可打开下面这行：
            # tqdm.write(f"[批次 {batch_idx}] ⏩ 跳过非JSON行 #{idx}: {line[:80]}")
            continue

        # 统计这一行的大括号数量
        open_cnt = line.count("{")
        close_cnt = line.count("}")

        if depth == 0:
            # 刚刚进入一个新的对象
            cur_lines = [line]
            depth = open_cnt - close_cnt

            # depth <= 0 说明这是“单行对象”：{"a":1} 或 {"a":1},
            if depth <= 0:
                text = "\n".join(cur_lines)
                text = _normalize_json_text(text)
                try:
                    obj = json.loads(text)
                except JSONDecodeError as e:
                    tqdm.write(f"[批次 {batch_idx}] ❌ JSON解析失败（行#{idx}附近）：{e}")
                    tqdm.write(f"[批次 {batch_idx}] 该对象原文：{text}")
                else:
                    # 统一 "话题簇1"/"话题簇2"... -> "话题簇"
                    for key in list(obj.keys()):
                        if key.startswith("话题簇") and key != "话题簇":
                            obj["话题簇"] = obj.pop(key)
                    objs.append(obj)
                cur_lines = []
                depth = 0
        else:
            # 已经在一个 JSON 对象内部（多行场景）
            cur_lines.append(line)
            depth += open_cnt - close_cnt

            # depth 回到 0，表示一个对象结束
            if depth <= 0:
                text = "\n".join(cur_lines)
                text = _normalize_json_text(text)

                try:
                    obj = json.loads(text)
                except JSONDecodeError as e:
                    tqdm.write(f"[批次 {batch_idx}] ❌ JSON解析失败（行#{idx}附近）：{e}")
                    tqdm.write(f"[批次 {batch_idx}] 该对象原文：{text}")
                else:
                    for key in list(obj.keys()):
                        if key.startswith("话题簇") and key != "话题簇":
                            obj["话题簇"] = obj.pop(key)
                    objs.append(obj)

                cur_lines = []
                depth = 0

    return objs
    
def parse_jsonl_text_safe(text: str, label: str = "模型#3聚合输出") -> List[Dict[str, Any]]:
    """
    尝试把 text 按行解析为 JSON 对象：
    - 每行独立尝试 json.loads
    - 解析前先对“极轴”等常见错误做一次正则修复
    - 解析失败时不会抛异常，而是打印出具体行号和原文，方便排查
    """
    objs: List[Dict[str, Any]] = []
    lines = text.strip().splitlines()

    for idx, raw in enumerate(lines, start=1):
        s = (raw or "").strip()
        if not s:
            continue

        # 跳过 ```json / ``` 等代码块标记
        if s.startswith("```"):
            continue

        # ⭐ 先修复“极轴”相关格式错误
        s_fixed = fix_model3_line_extreme_axis(s)

        try:
            obj = json.loads(s_fixed)
        except JSONDecodeError as e:
            print(f"[{label}] ❌ JSON解析失败：行#{idx} -> {e}")
            print(f"[{label}] 该行原文(修复后)：{s_fixed}")
            continue

        objs.append(obj)

    return objs

def fix_model3_line_extreme_axis(s: str) -> str:
    """
    专门修复模型#3输出中“极轴”相关的坏格式：
    针对以下几种情况：
    1) "日期": "极轴": "2025-12-06"  =>  "日期": "2025-12-06"
    2) "时间轴": "极轴": "22:30:57-22:33:57"  =>  "时间轴": "22:30:57-22:33:57"
    3) "时间轴": "22:34:24-极轴": "22:38:33" => "时间轴": "22:34:24-22:38:33"
    """

    # 1) 修复日期字段被“极轴”污染：
    #    "日期": "极轴": "2025-12-06"
    s = re.sub(
        r'"日期"\s*:\s*"极轴"\s*:\s*"([^"]+)"',
        r'"日期": "\1"',
        s
    )

    # 2) 修复时间轴字段完全是 "极轴": "xxx" 的情况：
    #    "时间轴": "极轴": "22:30:57-22:33:57"
    s = re.sub(
        r'"时间轴"\s*:\s*"极轴"\s*:\s*"([^"]+)"',
        r'"时间轴": "\1"',
        s
    )

    # 3) 修复时间轴值中间带 "-极轴": " 的情况：
    #    "时间轴": "22:34:24-极轴": "22:38:33"
    #    -> "时间轴": "22:34:24-22:38:33"
    s = re.sub(
        r'"时间轴"\s*:\s*"([^"]*?)-极轴"\s*:\s*"([^"]+)"',
        r'"时间轴": "\1-\2"',
        s
    )

    return s


def _strip_fences(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"^```[a-zA-Z0-9]*\s*", "", s)
    s = re.sub(r"\s*```$", "", s)
    return s.strip()

def _as_list(v):
    if v is None:
        return []
    if isinstance(v, list):
        return v
    return [v]

def _merge_time_axes(time_axes: list[str]) -> str:
    seen, out = set(), []
    for t in time_axes:
        t = (t or "").strip()
        if not t or t in seen:
            continue
        seen.add(t)
        out.append(t)
    return "、".join(out)


def clean_time_axis(raw: str) -> str:
    """
    保留：数字、冒号、横杠、顿号、空白；其他全部删掉
    """
    if not raw:
        return ""
    return re.sub(r"[^\d:、\-\s]", "", str(raw))


def normalize_model3_clusters(output_text: str, parsed_subclusters: list[dict]):
     # 0) 先让子簇具备 日期/时间轴
    parsed_subclusters = enrich_subclusters_with_datetime(parsed_subclusters)
   
    

    # 1) 建索引：_cluster_id -> 子簇(含日期/时间轴)
    sub_by_id = {}
    for sc in parsed_subclusters:
        cid = (sc.get("_cluster_id") or "").strip()
        if cid:
            sub_by_id[cid] = sc

    # 2) 逐行修复“极轴”污染再 parse
    raw = _strip_fences(output_text)
    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    lines = [fix_model3_line_extreme_axis(ln) for ln in lines]

    objs = []
    for ln in lines:
        try:
            objs.append(json.loads(ln))
        except Exception:
            continue

    normalized = []
    for obj in objs:
        topic = (obj.get("话题簇") or obj.get("聚合话题簇") or obj.get("话题簇3") or "").strip()

        sub_list = (
            obj.get("子话题簇列表")
            or obj.get("子话题簇")
            or obj.get("子话题簇id列表")
            or obj.get("子话题簇ID列表")
            or []
        )
        sub_list = [str(x).strip() for x in _as_list(sub_list) if str(x).strip()]

        date = (obj.get("日期") or "").strip()
        time_axis = (obj.get("时间轴") or "").strip()

        # 3) 缺失回填：从子簇取日期/时间轴
        if (not date) or (not time_axis):
            sub_dates, sub_axes = [], []
            for cid in sub_list:
                sc = sub_by_id.get(cid)
                if not sc:
                    continue
                d = (sc.get("日期") or "").strip()
                ta = (sc.get("时间轴") or "").strip()
                if d:
                    sub_dates.append(d)
                if ta:
                    sub_axes.append(ta)
            if not date and sub_dates:
                date = sub_dates[0]
            if not time_axis and sub_axes:
                time_axis = _merge_time_axes(sub_axes)

        # 4) 强制 schema：缺关键字段就丢弃，避免后续 split(None) 崩
        if not topic or not sub_list or not date or not time_axis:
            continue

        normalized.append({
            "话题簇": topic,
            "子话题簇列表": sub_list,
            "日期": date,
            "时间轴": clean_time_axis(time_axis)
        })

    return normalized, parsed_subclusters

######匹配all_CLUSTER时间

# 匹配：……（2025-12-07 22:40:36-22:41:27） 或 ……(2025-12-07 22:40:36-22:41:27)

SUB_TIME_RE = re.compile(
    r"[（(]?\s*(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2}-\d{2}:\d{2}:\d{2})\s*[）)]?"
)



def enrich_subclusters_with_datetime(parsed_subclusters: list[dict]) -> list[dict]:
    out = []
    for sc in parsed_subclusters:
        sc2 = dict(sc)
        title = sc2.get("话题簇") or ""
        m = SUB_TIME_RE.search(title)

        if m:
            sc2["日期"] = m.group(1)
            sc2["时间轴"] = m.group(2)
        else:
            # ✅ 新增：从“发言时间”里补齐
            ft = (sc2.get("发言时间") or "").strip()
            m2 = SUB_TIME_RE.search(ft)
            if m2:
                sc2["日期"] = m2.group(1)
                sc2["时间轴"] = m2.group(2)
            else:
                # 兜底：日期从 _cluster_id 取
                cid = (sc2.get("_cluster_id") or "").strip()
                if cid and re.match(r"^\d{4}-\d{2}-\d{2}_", cid):
                    sc2["日期"] = cid[:10]
                sc2.setdefault("时间轴", "")

        out.append(sc2)
    return out

#################################话提簇唯一id#################################

###1. clusterid的日期扒取##
def infer_date_for_batch(
    cluster_json_list: List[Dict[str, Any]],
    batch_lines: List[str]
) -> str:
    """
    优先从话题簇标题中抽取日期（形如：xxxx（2025-12-02 16:30:01-17:42:41））
    若失败，则回退到原始发言中的 `发言日期` / `日期` 字段。
    """
    # 1）先从话题簇标题里找 YYYY-MM-DD
    for obj in cluster_json_list:
        title = str(
            obj.get("话题簇")
            or obj.get("聚合话题簇")
            or ""
        )
        m = re.search(r"(\d{4}-\d{2}-\d{2})", title)
        if m:
            return m.group(1)

    # 2）如果标题里没有，就从原始发言里拿
    for line in batch_lines:
        line = line.strip()
        if not line:
            continue
        try:
            msg = json.loads(line)
        except Exception:
            continue

        date_str = msg.get("发言日期") or msg.get("日期")
        if date_str:
            return date_str

    # 3）再不行就报错，让你意识到数据格式有问题
    raise ValueError("无法从话题簇标题或原始发言中推断出日期，请检查数据格式。")

##2.生成话提簇clusterid##
def assign_global_cluster_ids(cluster_list, date_str, batch_id):
    """
    为每个话题簇生成全局唯一ID字段 `_cluster_id`
    格式：YYYY-MM-DD_BX_XX，如 2025-11-20_B2_03
    """
    for idx, cluster in enumerate(cluster_list, start=1):
        cluster["_cluster_id"] = f"{date_str}_{batch_id}_{idx:02d}"
    return cluster_list

#################################聚合每天的话提簇分批输出#################################
def aggregate_cluster_outputs(batch_outputs: List[str]) -> str:
    all_lines = []

    for batch_id, text in enumerate(batch_outputs, start=1):
        if not text:
            continue

        for line in text.strip().splitlines():
            line = line.strip()
            if not line:
                continue

            try:
                obj = json.loads(line)
            except Exception:
                # 如果有一行不是合法 JSON，就跳过（也可以改成 raise，看你需求）
                continue

            clean_line = json.dumps(obj, ensure_ascii=False)
            all_lines.append(clean_line)

    # 聚合为一个大的 JSONL 字符串
    return "\n".join(all_lines)


################top5筛选#################################


# -------------------------------
# 🔹 1. 匹配原始发言
# -------------------------------
from datetime import datetime
import json
from typing import List, Dict

import json
from datetime import datetime
from typing import List, Tuple

from datetime import datetime
from typing import List, Tuple

import re
from datetime import datetime

def parse_time_range(date_str: str, range_str: str):
    """
    从 range_str 中鲁棒地解析出一个时间段：
    - 支持出现“极轴”“极”等脏字符
    - 只认里面的 HH:MM:SS 模式
    - 解析失败时返回 (None, None)，避免直接抛异常
    """
    if not range_str:
        return None, None

    # 提取所有形如 16:10:56 的时间片段
    times = re.findall(r"\d{1,2}:\d{2}:\d{2}", str(range_str))
    if len(times) < 2:
        print(f"[parse_time_range] 无法从 {range_str!r} 提取到 2 个时间，跳过")
        return None, None

    start_str, end_str = times[0], times[1]

    try:
        start_dt = datetime.strptime(f"{date_str} {start_str}", "%Y-%m-%d %H:%M:%S")
        end_dt   = datetime.strptime(f"{date_str} {end_str}", "%Y-%m-%d %H:%M:%S")
        return start_dt, end_dt
    except ValueError as e:
        print(f"[parse_time_range] 解析失败：date_str={date_str!r}, range_str={range_str!r}, err={e}")
        return None, None


from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple


def match_dialogs_by_time(
    messages: List[Dict[str, Any]],
    date_str: str,
    time_axis_str: Optional[str],
) -> List[Dict[str, Any]]:
    """
    根据 日期 + 时间轴，从 messages 中筛选出对应的原始发言：
    - messages 里的时间字段为：发言日期(YYYY-MM-DD) + 发言时间(HH:MM:SS)
    - time_axis_str 支持多段："16:10:56-16:23:00、21:00:00-21:10:00"
    - 内部用 parse_time_range 做“极轴”鲁棒解析
    """
    if not messages or not date_str:
        return []
    if not time_axis_str or not isinstance(time_axis_str, str) or not time_axis_str.strip():
        return []

    # 解析所有时间段 -> List[(start_dt, end_dt)]
    ranges: List[Tuple[datetime, datetime]] = []
    for part in str(time_axis_str).split("、"):
        part = part.strip()
        if not part:
            continue
        start_dt, end_dt = parse_time_range(date_str, part)
        if start_dt and end_dt:
            ranges.append((start_dt, end_dt))

    if not ranges:
        return []

    matched: List[Dict[str, Any]] = []
    seen = set()

    for msg in messages:
        if (msg.get("发言日期") or "") != date_str:
            continue

        ts = msg.get("发言时间") or ""
        try:
            msg_dt = datetime.strptime(f"{date_str} {ts}", "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue

        hit = any(start <= msg_dt <= end for (start, end) in ranges)
        if not hit:
            continue

        key = (
            msg.get("_idx"),
            msg.get("发言日期"),
            msg.get("发言时间"),
            msg.get("玩家ID"),
            msg.get("玩家消息"),
        )
        if key in seen:
            continue
        seen.add(key)
        matched.append(msg)

    return matched

from datetime import datetime
from typing import List, Tuple

def _parse_time_ranges(fayan_time: str) -> List[Tuple[datetime.time, datetime.time]]:
    """
    把 '14:00:01-14:09:52、21:00:00-21:10:00' 解析成 List[(start_time, end_time)]
    只返回 time，不返回 datetime，避免和下游比较类型不一致。
    """
    if not fayan_time:
        return []

    parts = [p.strip() for p in str(fayan_time).split("、") if p.strip()]
    ranges: List[Tuple[datetime.time, datetime.time]] = []

    for part in parts:
        # 提取所有形如 16:10:56 的时间片段（兼容“极轴”等脏字符）
        times = re.findall(r"\d{1,2}:\d{2}:\d{2}", part)
        if len(times) < 2:
            continue
        start_str, end_str = times[0], times[1]
        try:
            start_t = datetime.strptime(start_str, "%H:%M:%S").time()
            end_t   = datetime.strptime(end_str,   "%H:%M:%S").time()
            ranges.append((start_t, end_t))
        except Exception:
            continue

    return ranges



def get_dialogs_lines_by_fayan_time_debug(
    jsonl_lines01: list[str],
    date_str: str,
    fayan_time: str,
    debug: bool = True,
) -> list[str]:
    """
    debug=True 且结果为空时，会打印：
    - 传入的 date/time_axis
    - 解析后的时间段
    - 当天原文最早/最晚时间
    - 当天原文总条数
    - 前几条原文时间样例
    """
    if not date_str:
        if debug:
            print("🧯[DEBUG] date_str 为空，无法回溯")
        return []

    time_ranges = _parse_time_ranges(fayan_time)
    if not time_ranges:
        if debug:
            print("🧯[DEBUG] 时间轴解析失败")
            print("  date_str =", date_str)
            print("  fayan_time =", repr(fayan_time))
        return []

    # 先扫描当天的所有时间，得到 min/max（用于判断是不是日期错了 or 原文时间格式不对）
    day_times = []
    day_samples = []
    for line in jsonl_lines01:
        s = (line or "").strip()
        if not s:
            continue
        try:
            obj = json.loads(s)
        except json.JSONDecodeError:
            continue
        if (obj.get("发言日期") or "") != date_str:
            continue
        ts = obj.get("发言时间") or ""
        try:
            t = datetime.strptime(ts, "%H:%M:%S").time()
        except Exception:
            continue
        day_times.append(t)
        if len(day_samples) < 8:
            day_samples.append(ts)

    # 正式过滤
    out = []
    for line in jsonl_lines01:
        s = (line or "").strip()
        if not s:
            continue
        try:
            obj = json.loads(s)
        except json.JSONDecodeError:
            continue

        if (obj.get("发言日期") or "") != date_str:
            continue

        ts = obj.get("发言时间") or ""
        try:
            t = datetime.strptime(ts, "%H:%M:%S").time()
        except Exception:
            continue

        for start_t, end_t in time_ranges:
            if start_t <= t <= end_t:
                out.append(s)
                break

    # 结果为空 -> 打印锁定信息
    if debug and not out:
        print("\n" + "🧯"*20)
        print("🧯[DEBUG] 时间段内没有原文，锁定信息如下：")
        print("  date_str =", date_str)
        print("  fayan_time =", fayan_time)
        print("  parsed_ranges =", [(a.strftime('%H:%M:%S'), b.strftime('%H:%M:%S')) for a,b in time_ranges])

        if day_times:
            print("  day_count =", len(day_times))
            print("  day_min =", min(day_times).strftime("%H:%M:%S"))
            print("  day_max =", max(day_times).strftime("%H:%M:%S"))
            print("  day_time_samples =", day_samples)
        else:
            print("  day_count = 0  👉 说明：这个 date_str 在原文里根本不存在（日期不一致）")

        # 额外：把目标时间段打印出来，检查是否超出当天范围
        for a, b in time_ranges:
            if day_times and (b < min(day_times) or a > max(day_times)):
                print("  ⚠ 目标时间段完全落在当天原文范围之外（很可能：日期错 or 时间轴错）")
                break

        print("🧯"*20 + "\n")

    return out


def get_time_axis(cluster: dict) -> str | None:
    # 优先标准字段
    for key in ("时间轴", "time轴", "时间轴 ", "时间段", "时间区间"):
        val = cluster.get(key)
        if isinstance(val, str) and val.strip():
            return val.strip()
    return None

def extract_cluster_stats(聚合话题簇列表: List[Dict], 原始发言: List[str]) -> List[Dict]:
    parsed_msgs = [json.loads(line.strip()) for line in 原始发言 if line.strip()]
    results = []

    for cluster in 聚合话题簇列表:
        date = cluster.get("日期")
        time_axis = get_time_axis(cluster)
        if not date or not time_axis:
            print(f"⚠ 聚合话题簇缺少日期或时间轴，跳过：{cluster}")
            continue
        matched = match_dialogs_by_time(parsed_msgs, date, time_axis)
        players = {msg.get("玩家ID") for msg in matched if msg.get("玩家ID")}
        result = {
            "聚合话题簇": cluster.get("话题簇") or cluster.get("聚合话题簇"),
            "子话题簇列表": cluster.get("子话题簇列表"),
            "发言玩家总数": len(players),
            "发言总数": len(matched)
        }
        results.append(result)

    return results

# -------------------------------
# 🔹 2. 热度识别主函数
# -------------------------------
def compute_heat_score(U: int, M: int) -> float:
    if U == 0 or M == 0:
        return 0.0
    return round(U * math.sqrt(M), 2)

def extract_top5_heat_clusters(聚合话题簇列表: List[Dict], 原始发言: List[str], top_k=5) -> List[Dict]:
    parsed_msgs = [json.loads(line.strip()) for line in 原始发言 if line.strip()]
    enriched = []

    for cluster in 聚合话题簇列表:
        date = cluster.get("日期")
        time_axis = cluster.get("时间轴")
        matched = match_dialogs_by_time(parsed_msgs, date, time_axis)

        players = {msg.get("玩家ID") for msg in matched if msg.get("玩家ID")}
        U = len(players)
        M = len(matched)
        heat = compute_heat_score(U, M)

        enriched.append({
        "聚合话题簇": cluster.get("话题簇") or cluster.get("聚合话题簇") or "未知",
        "子话题簇列表": cluster.get("子话题簇列表"),
        "日期": date,
        "时间轴": time_axis,
        "发言玩家总数": U,
        "发言总数": M,
        "热度评分": heat
         })


    # 按热度评分排序，取 TopK
    enriched.sort(key=lambda x: x["热度评分"], reverse=True)
    return enriched[:top_k]

# -------------------------------
# 🔹 3. 添加讨论点字段
# -------------------------------
def attach_discussion_points_day(top_clusters: List[Dict], subclusters: List[Dict]) -> List[Dict]:
    """
    将 top_clusters 中每个聚合簇的子话题簇列表，与 subclusters 中的 _cluster_id 匹配，
    拼出核心机制描述，并格式化为列点 + 空行。移除子话题簇列表字段。
    """
    # 构建 _cluster_id → 核心机制 映射
    cluster_mechanism_map = {
        row["_cluster_id"]: row["核心对象/机制"]
        for row in subclusters
        if "_cluster_id" in row and "核心对象/机制" in row
    }

    result = []
    for cluster in top_clusters:
        ids = cluster.get("子话题簇列表", [])
        mechanisms = [cluster_mechanism_map.get(cid) for cid in ids if cid in cluster_mechanism_map]

        # 格式化为列点 + 空行
        discussion_point = [m for m in mechanisms if m]

        # 修复聚合话题簇名称字段为空的问题
        cluster_name = cluster.get("话题簇") or cluster.get("聚合话题簇") or "未知"

        enriched_cluster = {
            "聚合话题簇": cluster_name,
            "日期": cluster.get("日期"),
            "时间轴": cluster.get("时间轴"),
            "发言玩家总数": cluster.get("发言玩家总数"),
            "发言总数": cluster.get("发言总数"),
            "热度评分": cluster.get("热度评分"),
            "讨论点": discussion_point
        }

        result.append(enriched_cluster)

    return result


def attach_discussion_points_all(
    top_clusters: List[Dict[str, Any]],
    subclusters: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    对每个聚合簇：
    - 根据 子话题簇列表(_cluster_id) 找回子话题簇
    - 用 子话题簇["核心对象/机制"]（或["讨论点"]） 作为讨论点文本
    - 只输出：讨论点 + 日期时间轴列表 + 子话题簇列表
    不做 TopK，不做热度，不排序
    """

    # _cluster_id -> 子话题簇row
    sub_by_id: Dict[str, Dict[str, Any]] = {}
    for row in subclusters:
        cid = row.get("_cluster_id")
        if cid:
            sub_by_id[str(cid)] = row

    result: List[Dict[str, Any]] = []

    for cluster in top_clusters:
        ids = cluster.get("子话题簇列表", []) or []

        # 用 point_text 聚合：同一个讨论点可能对应多个子簇
        point_bucket: Dict[str, Dict[str, Any]] = {}

        for cid in ids:
            sc = sub_by_id.get(str(cid))
            if not sc:
                continue

            point_text = (sc.get("核心对象/机制") or sc.get("讨论点") or "").strip()
            if not point_text:
                continue

            d = sc.get("日期") or cluster.get("日期")
            t_axis = sc.get("时间轴") or sc.get("时间范围") or sc.get("日期时间轴") or ""

            if point_text not in point_bucket:
                point_bucket[point_text] = {
                    "讨论点": point_text,
                    "日期时间轴列表": [],
                    "子话题簇列表": [],
                }

            point_bucket[point_text]["子话题簇列表"].append(str(cid))
            point_bucket[point_text]["日期时间轴列表"].append({
                "日期": d,
                "时间轴": t_axis
            })

        # 不排序、不截断：全部返回
        points = list(point_bucket.values())

        cluster_name = cluster.get("话题簇") or cluster.get("聚合话题簇") or "未知"

        enriched_cluster = {
            "聚合话题簇": cluster_name,
            "日期": cluster.get("日期"),
            "时间轴": cluster.get("时间轴"),
            "发言玩家总数": cluster.get("发言玩家总数"),
            "发言总数": cluster.get("发言总数"),
            "热度评分": cluster.get("热度评分"),
            "讨论点": points,
        }

        result.append(enriched_cluster)

    return result

####################### 存入每日发言 top5 ########################


def append_daily_top5_to_version_jsonl(
    final_result: List[Dict],
    version_jsonl_path: str = "version_daily_top5_with_opinion.jsonl",
):
    """
    将当日 Top5 追加写入某个【版本累计 jsonl 文件】中。
    
    同时为每条记录补充两个字段：
    1）_idx：在 version_jsonl_path 中的全局递增行号（从1开始）
    2）_daily_top_id：当天内的 Top 话题簇ID，形式：YYYY-MM-DD_TXX，例如 2025-12-03_T01
    
    说明：
    - 假设 final_result 内所有记录的 "日期" 相同（即同一天的 top5）
    - 如果文件中已存在同一日期的数据，会在原有基础上继续累加 _daily_top_id 的编号
    - 不会修改任何已有的 `_cluster_id` 字段（子话题簇仍然用它）
    """
    if not final_result:
        print("⚠ final_result 为空，今日无 Top5 可写入。")
        return

    # 取当天日期（假设 final_result 同一批都是同一天）
    date_str = final_result[0].get("日期")
    if not date_str:
        raise ValueError("final_result 中缺少 '日期' 字段，无法生成 _daily_top_id。")

    path = Path(version_jsonl_path)

    # ---------- 1. 计算当前文件中的最大 _idx（全局行号） ----------
    global_max_idx = 0
    # 同时统计“今天”已经有多少条，用于 _daily_top_id 续编号
    existing_count_for_day = 0

    if path.exists():
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError:
                    continue

                # 统计全局 _idx
                if isinstance(obj.get("_idx"), int):
                    global_max_idx = max(global_max_idx, obj["_idx"])
                else:
                    # 旧数据没有 _idx，就简单当作一行，新行号往后推
                    global_max_idx += 1

                # 统计当天的条数（用于 _daily_top_id）
                if obj.get("日期") == date_str:
                    existing_count_for_day += 1

    # ---------- 2. 追加写入，并为新记录生成 _idx + _daily_top_id ----------
    mode = "a" if path.exists() else "w"
    with path.open(mode, encoding="utf-8") as f:
        for offset, row in enumerate(final_result, start=1):
            row = dict(row)  # 避免修改原对象引用

            # 2.1 全局递增 _idx（版本所有天共用一个序号）
            global_max_idx += 1
            row["_idx"] = global_max_idx

            # 2.2 生成当天的 top id（不碰 `_cluster_id`）
            # 当天已有 existing_count_for_day 条，这批从 +1 开始接着排
            idx_for_day = existing_count_for_day + offset
            row["_daily_top_id"] = f"{date_str}_T{idx_for_day:02d}"

            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    print(f"✅ 已将当日 Top5（含 _idx 和 _daily_top_id）追加写入: {path}")


#########匹配final+output ==>讨论点+时间轴################
import json
import re
from typing import List, Dict, Any

def parse_jsonl_text(text: str) -> List[Dict[str, Any]]:
    """
    解析模型输出的“伪 jsonl”，尽量从中提取出若干合法的 dict。
    - 支持开头/结尾有 ```json 代码块包裹
    - 跳过非 { 开头的行
    - 自动去掉行尾的逗号
    - 解析失败时打印 warning，不中断
    """
    if not text:
        return []

    s = text.strip()

    # 去掉 ```json / ``` 包裹
    if s.startswith("```"):
        s = re.sub(r"^```[a-zA-Z0-9]*\s*", "", s)
        s = re.sub(r"\s*```$", "", s)

    results: List[Dict[str, Any]] = []

    for raw in s.splitlines():
        line = raw.strip()
        if not line:
            continue

        # 有些模型会输出 "- {...}" 之类的 markdown 列表
        if line.startswith("- "):
            line = line[2:].lstrip()

        # 跳过明显不是 JSON 对象的行
        if not line.startswith("{"):
            continue

        # 去掉末尾多余的逗号
        if line.endswith(","):
            line = line[:-1].rstrip()

        try:
            obj = json.loads(line)
        except json.JSONDecodeError as e:
            print(f"⚠ 解析 JSON 行失败：{e} | 行内容前 120 字符：{line[:120]}")
            continue

        if isinstance(obj, dict):
            results.append(obj)
        else:
            print(f"⚠ 解析结果不是 dict，已跳过：{type(obj)}")

    return results



def print_mech_time_from_top5(top5_results: list[dict], output_cluster_jsonl: str):
    output_clusters = parse_jsonl_text(output_cluster_jsonl)

    # 核心对象/机制 -> 发言时间
    mech_time = {}
    for r in output_clusters:
        mech = (r.get("核心对象/机制") or "").strip()
        t = (r.get("发言时间") or "").strip()
        if mech and mech not in mech_time:
            mech_time[mech] = t

    rows = []
    miss = 0

    for top in top5_results:
        for dp in (top.get("讨论点") or []):
            mech = (dp or "").strip()
            if not mech:
                continue

            row = {
                "核心对象/机制": mech,
                "发言时间": mech_time.get(mech, "")
            }
            rows.append(row)

            # 打印 jsonl 行
            #print(json.dumps(row, ensure_ascii=False))

            if not row["发言时间"]:
                miss += 1

    # 额外给你一个小提示，方便确认有没有漏匹配
    if miss:
        print(f"[WARN] 有 {miss} 条讨论点未匹配到发言时间（请检查是否完全一致）")

    return rows

######################锁定时间还原原文#############

def parse_fayan_time_range_str(fayan_time: str):
    """
    输入:  '2025-12-04 14:00:01-14:09:52'
    输出:  ('2025-12-04', '14:00:01', '14:09:52')
    """
    if not fayan_time:
        return None
    s = str(fayan_time).strip()
    m = re.search(r"(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2})\s*-\s*(\d{2}:\d{2}:\d{2})", s)
    if not m:
        return None
    return m.group(1), m.group(2), m.group(3)

def get_dialogs_lines_by_fayan_time(jsonl_lines01: list[str], fayan_time: str) -> list[str]:
    """
    从原始 jsonl_lines01 (list[str]) 里筛出落在发言时间范围内的所有原文行（仍然返回 list[str]）。
    """
    parsed = parse_fayan_time_range_str(fayan_time)
    if not parsed:
        return []
    date_str, start_str, end_str = parsed

    start_t = datetime.strptime(start_str, "%H:%M:%S").time()
    end_t   = datetime.strptime(end_str, "%H:%M:%S").time()

    out = []
    for line in jsonl_lines01:
        s = (line or "").strip()
        if not s:
            continue
        try:
            obj = json.loads(s)
        except json.JSONDecodeError:
            continue

        if (obj.get("发言日期") or "") != date_str:
            continue

        ts = obj.get("发言时间") or ""
        try:
            t = datetime.strptime(ts, "%H:%M:%S").time()
        except Exception:
            continue

        if start_t <= t <= end_t:
            out.append(s)  # ✅ 保持原样：一条json字符串
    return out

#------------------------模型4输出拆分---------------------------------

def parse_opinion_output_to_list(opinion_output: str) -> List[Dict[str, Any]]:
    """
    把模型4返回的字符串解析成 List[dict]：
    - 支持：
      * 单个 JSON 对象
      * JSON 数组
      * jsonl（多行，每行一个 JSON）
      * ```json 代码块 + 多行漂亮 JSON
    """
    if not opinion_output:
        return []

    s = opinion_output.strip()

    # 0) 去掉 ```json / ``` 外壳
    s = re.sub(r"^```[a-zA-Z0-9]*\s*", "", s)
    s = re.sub(r"\s*```$", "", s)
    s = s.strip()

    # 1) 先尝试整体解析（能吃掉单对象 / 数组）
    try:
        obj = json.loads(s)
        if isinstance(obj, dict):
            return [obj]
        if isinstance(obj, list):
            return [x for x in obj if isinstance(x, dict)]
    except json.JSONDecodeError:
        pass

    # 2) 用“括号深度”方式，从多行里拼出一个或多个 JSON 对象
    objs: List[Dict[str, Any]] = []
    buf: List[str] = []
    depth = 0

    for raw in s.splitlines():
        line = (raw or "").strip()
        if not line:
            continue

        # 去掉 markdown 列表前缀，如 "- { ... }"
        if line.startswith("- "):
            line = line[2:].lstrip()

        # 如果这一行完全不含 { 或 }，且当前深度为 0，基本是解释文字，跳过
        if depth == 0 and "{" not in line:
            continue

        # 进入/继续一个对象
        open_cnt = line.count("{")
        close_cnt = line.count("}")

        if depth == 0 and "{" in line:
            # 新对象的开始
            buf = [line]
            depth = open_cnt - close_cnt
            if depth <= 0:
                # 单行对象：{...}
                text = "\n".join(buf)
                try:
                    obj = json.loads(text)
                    if isinstance(obj, dict):
                        objs.append(obj)
                except json.JSONDecodeError:
                    pass
                buf = []
                depth = 0
        else:
            # 已经在对象内部
            buf.append(line)
            depth += open_cnt - close_cnt
            if depth <= 0:
                # 对象结束
                text = "\n".join(buf)
                try:
                    obj = json.loads(text)
                    if isinstance(obj, dict):
                        objs.append(obj)
                except json.JSONDecodeError:
                    pass
                buf = []
                depth = 0

    return objs

######################观点回溯至top5做最后输出##############
def _norm_text(s: str) -> str:
    """文本归一化：去空格和常见标点，用于宽松匹配讨论点."""
    if not s:
        return ""
    s = str(s)
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[，,。\.！？!：:、；;（）()\[\]【】\"'“”‘’]", "", s)
    return s


def merge_top5_with_opinions_numbered(
    top5_results: List[Dict[str, Any]],
    opinions: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:

    # 1) opinions -> multimap：归一化讨论点 => [op1, op2, ...]
    op_map: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for op in opinions:
        dp = (op.get("讨论点") or "").strip()
        if not dp:
            continue
        op_map[_norm_text(dp)].append(op)

    merged: List[Dict[str, Any]] = []
    missing = 0
    hit = 0

    # 2) 合并回 top5（逐个消费 list，避免覆盖丢失）
    for row in top5_results:
        dps = row.get("讨论点") or []
        if isinstance(dps, str):
            dps = [dps]

        discussion_list: List[Dict[str, Any]] = []

        for idx, dp in enumerate(dps, start=1):
            raw_dp = (dp or "").strip()
            if not raw_dp:
                continue

            k = _norm_text(raw_dp)
            numbered_key = f"讨论点{idx}"

            if not op_map.get(k):
                discussion_list.append({
                    numbered_key: raw_dp,
                    "玩家观点": [],
                    "代表性玩家发言示例": [],
                    "_missing_opinion": True,   # 调试用
                })
                missing += 1
                continue

            # ✅ 关键：按顺序取出并消费，防止同名覆盖
            op = op_map[k].pop(0)
            hit += 1

            examples = op.get("代表性玩家发言示例") or []
            if not isinstance(examples, list):
                examples = [str(examples)]

            viewpoints = op.get("玩家观点") or []
            if not isinstance(viewpoints, list):
                viewpoints = [str(viewpoints)]

            discussion_list.append({
                numbered_key: raw_dp,
                "玩家观点": viewpoints,
                "代表性玩家发言示例": examples,
            })

        new_row = dict(row)
        new_row.pop("讨论点", None)
        new_row.pop("代表性玩家发言示例", None)
        new_row["讨论点列表"] = discussion_list
        merged.append(new_row)

    # 3) 可选：看有没有 opinion 没被用掉（一般也能帮助确认是否 key 冲突）
    leftover = sum(len(v) for v in op_map.values())
    print(f"✅ merge 命中={hit}, missing={missing}, leftover_unused_opinions={leftover}")

    return merged


#####时间轴校正###
import re
from typing import Any, Dict, List, Optional, Tuple

def _pick_time_axis_value(c: Dict[str, Any]) -> Tuple[Optional[str], Optional[str]]:
    """
    在 dict c 中挑一个最像“时间轴”的字段值。
    返回 (value, source_key)
    """
    # 候选 key 的优先级：越靠前越像时间轴
    priority_patterns = [
        r"^时间轴$", r"时间轴",
        r"时间段", r"时间范围", r"时段", r"范围",
        r"time[_\s-]*axis", r"axis",
        r"极轴", r"时间极"
    ]

    # 1) 先按优先级找 key
    keys = list(c.keys())
    for pat in priority_patterns:
        for k in keys:
            if re.search(pat, str(k), flags=re.IGNORECASE):
                v = c.get(k)
                if isinstance(v, str) and v.strip():
                    return v.strip(), str(k)
                # 有些模型会给 list[str]，拼一下
                if isinstance(v, list) and v:
                    s = "、".join([str(x).strip() for x in v if str(x).strip()])
                    if s:
                        return s, str(k)

    # 2) 再扫所有 key：只要 key 含“时间/时段/范围/time/axis/极轴”
    for k in keys:
        kk = str(k)
        if any(x in kk.lower() for x in ["时间", "时段", "范围", "time", "axis", "极轴"]):
            v = c.get(k)
            if isinstance(v, str) and v.strip():
                return v.strip(), str(k)
            if isinstance(v, list) and v:
                s = "、".join([str(x).strip() for x in v if str(x).strip()])
                if s:
                    return s, str(k)

    return None, None


def ensure_time_axis_key(c: Dict[str, Any], verbose: bool = True) -> bool:
    """
    如果 c 中没有 key == '时间轴' 或其值为空，就用别名字段覆盖生成 c['时间轴']。
    返回：是否发生了修复（True/False）
    """
    # 已经有时间轴且非空 -> 不动
    ta = c.get("时间轴")
    if isinstance(ta, str) and ta.strip():
        return False

    # 没有“时间轴”三个字的 key（你要求：只要没出现时间轴三个字就覆盖）
    has_time_axis_key = any("时间轴" in str(k) for k in c.keys())
    if not has_time_axis_key:
        v, src = _pick_time_axis_value(c)
        if v:
            c["时间轴"] = v
            c["_time_axis_from"] = src  # 记录来源，方便排查
            if verbose:
                print(f"✅ 已补齐 时间轴: {v}  (from: {src}) | 聚合话题簇={c.get('聚合话题簇')}")
            return True
        else:
            if verbose:
                print(f"⚠ 找不到可用时间字段来补齐 时间轴 | 聚合话题簇={c.get('聚合话题簇')} keys={list(c.keys())}")
            return False

    # 有 key 含“时间轴”但值为空：也尝试用别名补齐
    v, src = _pick_time_axis_value(c)
    if v:
        c["时间轴"] = v
        c["_time_axis_from"] = src
        if verbose:
            print(f"✅ 已修复 空时间轴: {v} (from: {src}) | 聚合话题簇={c.get('聚合话题簇')}")
        return True

    if verbose:
        print(f"⚠ 有时间轴key但无法修复（无可用值）| 聚合话题簇={c.get('聚合话题簇')}")
    return False

def ensure_subcluster_list_key(c: Dict[str, Any]) -> bool:
    if isinstance(c.get("子话题簇列表"), list) and c["子话题簇列表"]:
        return False

    candidates = ["子话题簇列表", "子话提簇列表", "子话题簇ID列表", "子簇列表", "子话题簇"]
    for k in candidates:
        v = c.get(k)
        if v is None:
            continue
        if isinstance(v, list):
            c["子话题簇列表"] = v
            c["_sub_list_from"] = k
            return True
        if isinstance(v, str) and v.strip():
            c["子话题簇列表"] = [x.strip() for x in v.split("、") if x.strip()]
            c["_sub_list_from"] = k
            return True
    return False
#######按 _idx 回原文算真实时间轴 / 取原文行#######
from datetime import datetime
import json
import re

def extract_idx_list_from_cluster_obj(c: dict) -> list[int]:
    v = c.get("发言行号列表")
    if v is None:
        return []
    if isinstance(v, list):
        out = []
        for x in v:
            try:
                out.append(int(x))
            except:
                pass
        return out
    if isinstance(v, str):
        nums = re.findall(r"\d+", v)
        return [int(n) for n in nums]
    return []

def calc_fayan_time_by_idx(jsonl_lines01_idx: list[str], idx_list: list[int]) -> str:
    """
    用 idx_list 回原文算真实 min/max -> 'YYYY-MM-DD HH:MM:SS-HH:MM:SS'
    """
    if not idx_list:
        return ""
    idx_set = set(idx_list)

    dts = []
    for line in jsonl_lines01_idx:
        try:
            obj = json.loads(line)
        except:
            continue
        if obj.get("_idx") not in idx_set:
            continue

        d = (obj.get("发言日期") or "").strip()
        t = (obj.get("发言时间") or "").strip()
        if not d or not t:
            continue
        try:
            dt = datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M:%S")
        except:
            continue
        dts.append(dt)

    if not dts:
        return ""

    dts.sort()
    date_str = dts[0].strftime("%Y-%m-%d")
    return f"{date_str} {dts[0].strftime('%H:%M:%S')}-{dts[-1].strftime('%H:%M:%S')}"


def refill_cluster_fayan_time(cluster_json_list: list[dict], jsonl_lines01_idx: list[str]) -> int:
    ok = 0
    for c in cluster_json_list:
        idxs = extract_idx_list_from_cluster_obj(c)
        axis = calc_fayan_time_by_idx(jsonl_lines01_idx, idxs)
        c["发言时间"] = axis  # ✅ 写回：后面链路仍然按“发言时间”跑
        c["_发言时间来源"] = "idx_minmax" if axis else "idx_empty"
        ok += 1 if axis else 0
    return ok
