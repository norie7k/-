

from __future__ import annotations
import json, time, typing as T
import pandas as pd
import requests
from pathlib import Path

# --- openpyxl 样式/工具 ---
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


################模型调用，出结果###################

def load_system_prompt(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Prompt file not found: {path}")
    return path.read_text(encoding="utf-8")

def build_user_prompt_filter(json_lines: T.List[str]) -> str:
    # 模型#1：筛掉非游戏相关，只输出相关 JSON 行（原样）
    return (
        "以下是若干玩家/客服/研发的发言记录，请根据系统提示中规则，"
        "判断哪些是【与游戏内容相关】的发言，保留这些 JSON 行，不相关的忽略。"
        "请仅输出【相关发言的原始 JSON 行】，严格保持格式不变。\n\n"
        "【输入】：\n" + "\n".join(json_lines)
    )

def build_user_prompt_classify(jsonl_block: str) -> str:
    # 模型#2：对已筛选的相关 JSON 行（原样）进行分类，追加“意图分类”键
    return (
       "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请仅输出【 JSON 行】，\n\n"
        "【输入】：\n" + jsonl_block
    )
def build_user_prompt_classify2(jsonl_block: str) -> str:
    # 模型#3：对已筛选的相关 JSON 行（原样）进行二级标签分类
    return (
       "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请仅输出【 JSON 行】，\n\n"
        "【输入】：\n" + jsonl_block
    )

def call_ark_chat_completions(
    api_url: str,
    api_key: str,
    model: str,
    system_prompt: str,
    user_prompt: str,
    temperature: float = 0.2,
    max_tokens: int = 32700,
    timeout: int = 600,
    retries: int = 2,
) -> str:
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    last_err = None
    for attempt in range(retries + 1):
        try:
            resp = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text}")
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            last_err = e
            time.sleep(1.2 * (attempt + 1))
    raise RuntimeError(f"Ark API 调用失败: {last_err}")

def extract_valid_json_lines(text: str) -> T.List[str]:
    """
    把模型输出里的纯 JSON 行提取出来（鲁棒处理）：
    - 逐行判断：以 { 开头 且 以 } 结尾，则认为是一个 JSON 对象行
    - 也能容忍前后多余空行或解释文字（会被忽略）
    """
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("{") and s.endswith("}"):
            lines.append(s)
    return lines



import json
import pandas as pd

def jsonl_to_dataframe_with_intent(jsonl_text: str) -> pd.DataFrame:
    """
    将模型#3输出（JSONL，每行一个JSON）转 df。
    目标列：["发言时间","玩家ID","玩家消息","一级分类","二级标签"]
    """
    # 允许你已有的 extract_valid_json_lines，若没有就简单 splitlines
    try:
        pure_lines = extract_valid_json_lines(jsonl_text)
    except NameError:
        pure_lines = [ln.strip() for ln in (jsonl_text or "").splitlines() if ln.strip()]

    if not pure_lines:
        return pd.DataFrame(columns=["发言时间","玩家ID","玩家消息","一级分类","二级标签"])

    rows = []
    for line in pure_lines:
        try:
            rows.append(json.loads(line))
        except Exception:
            # 跳过模型可能输出的解释性文本
            continue

    if not rows:
        return pd.DataFrame(columns=["发言时间","玩家ID","玩家消息","一级分类","二级标签"])

    df = pd.DataFrame(rows)

    # 兼容键名：玩家 ID -> 玩家ID；意图分类/分类标签 -> 一级分类
    df = df.rename(columns={
        "玩家 ID": "玩家ID",
        "意图分类": "一级分类",
        "分类标签": "一级分类",
    })

    # 缺列补空
    for c in ["发言时间","玩家ID","玩家消息","一级分类","二级标签"]:
        if c not in df.columns:
            df[c] = pd.NA

    # 一级分类转数值（便于后续过滤/分发）
    df["一级分类"] = pd.to_numeric(df["一级分类"], errors="coerce")

    # 二级标签统一为 list[str]
    def _to_list(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return []
        if isinstance(x, list):
            return [str(i) for i in x]
        return [str(x)]
    df["二级标签"] = df["二级标签"].apply(_to_list)

    # 只保留目标列顺序
    return df[["发言时间","玩家ID","玩家消息","一级分类","二级标签"]]



##########################导入Excel格式要求###########################

from pathlib import Path
from typing import List, Dict, Any
import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

########################## Excel 结构与样式 ###########################
SHEET_NAMES = ["体验反馈", "疑惑询问", "建议灵感", "情绪输出", "问题反馈"]
# 新增“二级标签”列
HEADERS = ["发言时间", "玩家ID", "玩家消息", "二级标签"]

# 样式配置
FONT_NAME = "微雅软黑"         # 若本机没有，可改 "微软雅黑"
HEADER_FONT_SIZE = 12
BODY_FONT_SIZE   = 11
HEADER_FILL      = "DDDDDD"     # 灰底
HEADER_ROW_HEIGHT = 24
COL_WIDTHS = [21, 30, 95, 16]   # 时间 / 玩家ID / 玩家消息 / 二级标签

CA_TO_SHEET = {1:"体验反馈", 2:"疑惑询问", 3:"建议灵感", 4:"情绪输出", 5:"问题反馈"}


def _ensure_named_style(wb) -> str:
    """确保正文 NamedStyle 存在；返回 style 名称。"""
    style_name = "BodyStyle"
    if style_name in wb.named_styles:
        return style_name
    thin = Side(style="thin", color="000000")
    body_style = NamedStyle(name=style_name)
    body_style.font = Font(name=FONT_NAME, size=BODY_FONT_SIZE)
    body_style.alignment = Alignment(vertical="center", wrap_text=True)
    body_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.add_named_style(body_style)
    return style_name


def create_intent_excel_styled(filename: str):
    """初始化工作簿：5 个 sheet，表头+列宽+冻结+表头样式+正文样式注册"""
    wb = Workbook()
    ws0 = wb.active
    ws0.title = SHEET_NAMES[0]

    for idx, name in enumerate(SHEET_NAMES):
        ws = wb[name] if idx == 0 else wb.create_sheet(title=name)

        # 表头
        ws.append(HEADERS)

        # 列宽
        for col_idx, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # 冻结首行
        ws.freeze_panes = "A2"

        # 表头样式
        for c in range(1, len(HEADERS)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(name=FONT_NAME, size=HEADER_FONT_SIZE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)

        ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    _ensure_named_style(wb)
    wb.save(filename)
    print(f"✅ 创建并套样式：{filename}")


def _open_or_create_excel(excel_path: str):
    """没有文件就创建并预置格式；返回 (wb, created_flag)"""
    p = Path(excel_path)
    if not p.exists():
        create_intent_excel_styled(excel_path)
        wb = load_workbook(excel_path)
        return wb, True
    wb = load_workbook(excel_path)
    _ensure_named_style(wb)
    return wb, False


########################## 数据规范化 ###########################
def _normalize_records(records: List[Dict[str, Any]]) -> pd.DataFrame:
    """
    输入：形如
    [{"一级分类":2,"二级标签":["殖装保留","保留猜测"],"发言时间":"2025-08-06 17:25:36",
      "玩家 ID":"，(1272414483)","玩家消息":"能保留殖装应该"}, ...]
    输出：DataFrame，列 = ["一级分类","发言时间","玩家ID","玩家消息","二级标签"]，
    其中“二级标签”为单个字符串（若原本是列表则 explode）
    """
    if not records:
        return pd.DataFrame(columns=["一级分类", "发言时间", "玩家ID", "玩家消息", "二级标签"])

    df = pd.DataFrame(records)

    # 兼容 键名差异：玩家 ID vs 玩家ID
    if "玩家ID" not in df.columns and "玩家 ID" in df.columns:
        df = df.rename(columns={"玩家 ID": "玩家ID"})

    # 缺列补空
    for col in ["一级分类", "发言时间", "玩家ID", "玩家消息", "二级标签"]:
        if col not in df.columns:
            df[col] = pd.NA

    # 统一“二级标签”为列表
    def _to_list(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return []
        if isinstance(x, list):
            return x
        # 字符串就当成单标签
        return [str(x)]

    df["二级标签"] = df["二级标签"].apply(_to_list)

    # explode 成单标签一行；若原本为空列表，将生成空行，先过滤
    df = df.explode("二级标签", ignore_index=True)
    df["二级标签"] = df["二级标签"].fillna("").astype(str)

    # 一级分类数值化 & 过滤有效 sheet
    df["一级分类"] = pd.to_numeric(df["一级分类"], errors="coerce")
    df = df[df["一级分类"].isin(CA_TO_SHEET.keys())]

    # 只保留需要列、并按既定顺序
    return df[["一级分类", "发言时间", "玩家ID", "玩家消息", "二级标签"]]


########################## 写入 + 按二级标签热度排序 ###########################
def _rewrite_sheet_sorted_by_tag(wb, sheet_name: str, new_rows_df: pd.DataFrame):
    """
    将新的行与 sheet 既有行合并，然后按“二级标签”出现次数降序排序，重写正文区域。
    new_rows_df 列：["发言时间","玩家ID","玩家消息","二级标签"]
    """
    ws = wb[sheet_name]
    body_style_name = _ensure_named_style(wb)

    # 读取既有正文（A2:Dmax）
    existing = []
    if ws.max_row >= 2:
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row,
                              min_col=1, max_col=len(HEADERS), values_only=True):
            # 跳过全空行
            if r is None or all(x in (None, "") for x in r):
                continue
            existing.append(r)

    cols = HEADERS  # ["发言时间","玩家ID","玩家消息","二级标签"]
    df_exist = pd.DataFrame(existing, columns=cols) if existing else pd.DataFrame(columns=cols)

    # 合并
    df_all = pd.concat([df_exist, new_rows_df[cols]], ignore_index=True)

    if df_all.empty:
        # 清空正文
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        wb.save(excel_path)
        return

    # 计算标签热度并排序：先按热度(降序)，再按时间(升序)做稳定次序
    tag_counts = df_all["二级标签"].fillna("").astype(str).value_counts()
    df_all["_cnt"] = df_all["二级标签"].map(tag_counts)
    # 时间列可能为字符串，尽量解析；解析失败保持原字符串并次序由字符串比较决定
    try:
        dt = pd.to_datetime(df_all["发言时间"], errors="coerce")
        df_all["_ts"] = dt
        # 用 _ts 排序（None 放后面），再用原字符串兜底
        df_all = df_all.sort_values(by=["_cnt", "_ts", "发言时间"],
                                    ascending=[False, True, True])
    except Exception:
        df_all = df_all.sort_values(by=["_cnt", "发言时间"],
                                    ascending=[False, True])

    df_all = df_all.drop(columns=["_cnt"] + (["_ts"] if "_ts" in df_all.columns else []))

    # 清空正文区域
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # 逐行写回
    for row in df_all.itertuples(index=False, name=None):
        ws.append(row)

    # 套正文样式
    if ws.max_row >= 2:
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row,
                              min_col=1, max_col=len(HEADERS)):
            for cell in r:
                cell.style = body_style_name


def append_json_to_excel_by_cat_and_tag(records: List[Dict[str, Any]], excel_path: str):
    """
    主入口：
    - 解析/规范化输入记录
    - 按“一级分类”分发到 sheet
    - 每个 sheet 写入后，对“所有内容（含历史+新增）”按“二级标签”热度降序重排
    """
    df = _normalize_records(records)
    if df.empty:
        return

    wb, _ = _open_or_create_excel(excel_path)

    # 按一级分类分组，每个组在各自 sheet 内完成“合并 + 重排 + 重写”
    for ca, grp in df.groupby("一级分类", sort=False):
        sheet = CA_TO_SHEET.get(int(ca))
        if not sheet or sheet not in wb.sheetnames:
            continue

        # 仅保留写入列顺序
        body_df = grp[["发言时间", "玩家ID", "玩家消息", "二级标签"]].copy()

        _rewrite_sheet_sorted_by_tag(wb, sheet, body_df)

    wb.save(excel_path)


